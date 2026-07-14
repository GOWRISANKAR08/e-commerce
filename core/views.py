"""
Custom admin panel — the rich dashboard (faithful port of api/admin/stats +
admin/(dash)/page.tsx) plus management list pages for the catalog and orders.
Long-tail sections (banners, blogs, testimonials, faq, policies, enquiries,
reviews, users, offers) are managed through the built-in Django admin, linked
from the sidebar.
"""
import calendar
import csv
from datetime import datetime
from functools import wraps

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, Max, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from accounts.models import CustomerNote, User, UserType, Status
from core.nav import ADMIN_NAV
from core.utils import gen_code, slugify, next_position, insert_at_position, move_to_position, repack_positions
from store.models import (
    Banner, BannerType, Category, CMSRevision, ComboImage, ComboItem, ComboPackage,
    Coupon, CouponType, Enquiry, EnquiryReply, EnquiryStatus, Faq, FaqCategory,
    Notification, Order, OrderChannel, OrderEvent, OrderItem, OrderNote, OrderStatus,
    OrderRefund, PaymentMode, PaymentStatus,
    Policy, PolicyType, Product, ProductVariant, Review, ReviewStatus, SiteSettings,
    StockStatus, StockThreshold, StockStatusHistory, TeamMember, Testimonial,
)
from django.http import JsonResponse as _JsonResponse

# --------------------------------------------------------------------------- #
# Order workflow engine
# --------------------------------------------------------------------------- #
# Fixed progression — cannot skip or reverse.
ORDER_WORKFLOW = [
    OrderStatus.PROCESSING,
    OrderStatus.ORDER_CONFIRMED,
    OrderStatus.PACKED,
    OrderStatus.DISPATCHED,
    OrderStatus.DELIVERED,
]

# Valid next states from each status (including cancellation paths).
WORKFLOW_TRANSITIONS: dict[str, list[str]] = {
    OrderStatus.PROCESSING:     [OrderStatus.ORDER_CONFIRMED, OrderStatus.CANCELLED],
    OrderStatus.ORDER_CONFIRMED:[OrderStatus.PACKED,          OrderStatus.CANCELLED],
    OrderStatus.PACKED:         [OrderStatus.DISPATCHED,      OrderStatus.CANCELLED],
    OrderStatus.DISPATCHED:     [OrderStatus.DELIVERED,       OrderStatus.CANCELLED],
    OrderStatus.DELIVERED:      [OrderStatus.REFUNDED],
    OrderStatus.CANCELLED:      [],
    OrderStatus.REFUNDED:       [],
}

# Statuses where stock has already been deducted (confirmed or beyond).
STOCK_DEDUCTED_STATUSES = {
    OrderStatus.ORDER_CONFIRMED,
    OrderStatus.PACKED,
    OrderStatus.DISPATCHED,
    OrderStatus.DELIVERED,
}


def _valid_transition(from_status: str, to_status: str) -> bool:
    return to_status in WORKFLOW_TRANSITIONS.get(from_status, [])


def _deduct_stock_for_order(order: Order) -> None:
    """
    Deduct stock for every item in *order* (called on ORDER_CONFIRMED).
    Uses SELECT FOR UPDATE to be safe under concurrent requests.
    Prevents stock from going negative and auto-computes new status via thresholds.
    """
    with transaction.atomic():
        for item in order.items.select_related("variant__product").all():
            v = ProductVariant.objects.select_for_update().get(pk=item.variant_id)
            old_stock, old_status = v.stock, v.stock_status

            deducted = min(item.qty, v.stock)        # never below 0
            v.stock = max(v.stock - item.qty, 0)
            v.reserved_stock = v.reserved_stock + deducted

            threshold = StockThreshold.get_for_product(v.product)
            v.stock_status = threshold.compute_status(v.stock)
            v.save(update_fields=["stock", "reserved_stock", "stock_status"])

            if old_stock != v.stock or old_status != v.stock_status:
                StockStatusHistory.objects.create(
                    variant=v, old_status=old_status, new_status=v.stock_status,
                    old_stock=old_stock, new_stock=v.stock,
                    note=f"Order {order.order_id} confirmed",
                )


def _restore_stock_for_order(order: Order) -> None:
    """
    Restore reserved stock back to available stock (called on CANCELLED).
    Only acts when the order is in a state where stock was already deducted.
    """
    if order.status not in STOCK_DEDUCTED_STATUSES:
        return
    with transaction.atomic():
        for item in order.items.select_related("variant__product").all():
            v = ProductVariant.objects.select_for_update().get(pk=item.variant_id)
            old_stock, old_status = v.stock, v.stock_status

            restored = min(item.qty, v.reserved_stock)
            v.stock = v.stock + restored
            v.reserved_stock = max(v.reserved_stock - restored, 0)

            threshold = StockThreshold.get_for_product(v.product)
            v.stock_status = threshold.compute_status(v.stock)
            v.save(update_fields=["stock", "reserved_stock", "stock_status"])

            StockStatusHistory.objects.create(
                variant=v, old_status=old_status, new_status=v.stock_status,
                old_stock=old_stock, new_stock=v.stock,
                note=f"Order {order.order_id} cancelled — stock restored",
            )


def _release_reserved_on_deliver(order: Order) -> None:
    """
    Clear reserved_stock when an order is DELIVERED (goods have left the warehouse).
    Available stock is unchanged — it was already deducted at confirmation.
    """
    with transaction.atomic():
        for item in order.items.select_related("variant").all():
            v = ProductVariant.objects.select_for_update().get(pk=item.variant_id)
            v.reserved_stock = max(v.reserved_stock - item.qty, 0)
            v.save(update_fields=["reserved_stock"])


def admin_required(view):
    @wraps(view)
    @login_required(login_url="/admin-login")
    def wrapper(request, *args, **kwargs):
        if request.user.user_type != UserType.ADMIN:
            return redirect("/admin-login")
        return view(request, *args, **kwargs)
    return wrapper


def _ctx(active, **extra):
    base = {"ADMIN_NAV": ADMIN_NAV, "active_path": active}
    base.update(extra)
    return base


# --------------------------------------------------------------------------- #
# Dashboard — full stats port
# --------------------------------------------------------------------------- #
@admin_required
def dashboard(request):
    now = timezone.now()

    products = Product.objects.filter(status=Status.ACTIVE).count()
    orders = Order.objects.count()
    users = User.objects.filter(user_type=UserType.USER).count()
    reviews = Review.objects.count()
    enquiries = Enquiry.objects.filter(status=EnquiryStatus.OPEN).count()

    revenue = (Order.objects.filter(payment_status=PaymentStatus.PAID)
               .aggregate(s=Sum("grand_total"))["s"] or 0)
    avg_order = (Order.objects.filter(payment_status=PaymentStatus.PAID)
                 .aggregate(a=Avg("grand_total"))["a"] or 0)

    pipeline = {
        "pending": Order.objects.filter(status=OrderStatus.PROCESSING).count(),
        "confirmed": Order.objects.filter(status=OrderStatus.ORDER_CONFIRMED).count(),
        "packed": Order.objects.filter(status=OrderStatus.PACKED).count(),
        "dispatched": Order.objects.filter(status=OrderStatus.DISPATCHED).count(),
        "delivered": Order.objects.filter(status=OrderStatus.DELIVERED).count(),
        "returns": Order.objects.filter(
            status__in=[OrderStatus.CANCELLED, OrderStatus.REFUNDED]).count(),
    }

    _active_v = ProductVariant.objects.filter(product__status=Status.ACTIVE, status=Status.ACTIVE)
    inv = {
        "inStock":    _active_v.filter(stock_status=StockStatus.IN_STOCK).count(),
        "lowStock":   _active_v.filter(stock_status=StockStatus.LOW_STOCK).count(),
        "outOfStock": _active_v.filter(stock_status=StockStatus.OUT_OF_STOCK).count(),
        "total":      _active_v.count(),
    }
    inv["expiring"] = 0
    health_pct = round(inv["inStock"] / inv["total"] * 100) if inv["total"] else 0
    low_share = round(inv["lowStock"] / inv["total"] * 100) if inv["total"] else 0
    out_share = round(inv["outOfStock"] / inv["total"] * 100) if inv["total"] else 0

    recent_orders = (Order.objects.select_related("user")
                     .prefetch_related("items").order_by("-created_at")[:6])

    # 12-month revenue
    monthly = []
    for i in range(12):
        m = now.month - 11 + i
        y = now.year + (m - 1) // 12
        mm = (m - 1) % 12 + 1
        start = now.replace(year=y, month=mm, day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = calendar.monthrange(y, mm)[1]
        end = start.replace(day=last_day, hour=23, minute=59, second=59)
        rev = (Order.objects.filter(payment_status=PaymentStatus.PAID,
                                    created_at__gte=start, created_at__lte=end)
               .aggregate(s=Sum("grand_total"))["s"] or 0)
        monthly.append({"month": calendar.month_abbr[mm], "revenue": rev})

    max_rev = max([m["revenue"] for m in monthly] + [1])
    for m in monthly:
        m["pct"] = max(round(m["revenue"] / max_rev * 100), 4)
    avg_monthly = sum(m["revenue"] for m in monthly) / len(monthly) if monthly else 0
    best_month = max(monthly, key=lambda m: m["revenue"]) if monthly else None

    # top products by revenue
    top_items = (OrderItem.objects.values("product_id")
                 .annotate(revenue=Sum("net_total"), qty=Sum("qty"))
                 .order_by("-revenue")[:5])
    pid_map = {p.id: p for p in Product.objects.filter(
        id__in=[t["product_id"] for t in top_items]).select_related("category")}
    top_max = max([t["revenue"] or 0 for t in top_items] + [1])
    top_products = []
    for t in top_items:
        p = pid_map.get(t["product_id"])
        top_products.append({
            "product": p, "revenue": t["revenue"] or 0, "qty": t["qty"] or 0,
            "bar_pct": round((t["revenue"] or 0) / top_max * 100),
            "in_stock": (p.variants.first().stock_status != StockStatus.OUT_OF_STOCK
                         if p and p.variants.exists() else True),
        })

    low_stock_items = (_active_v.select_related("product")
                       .filter(stock_status__in=[StockStatus.LOW_STOCK, StockStatus.OUT_OF_STOCK])
                       .order_by("stock")[:4])

    # Inventory analytics for dashboard — active products/variants only
    inv_analytics = {
        "total_skus":  _active_v.count(),
        "in_stock":    _active_v.filter(stock_status=StockStatus.IN_STOCK).count(),
        "low_stock":   _active_v.filter(stock_status=StockStatus.LOW_STOCK).count(),
        "out_stock":   _active_v.filter(stock_status=StockStatus.OUT_OF_STOCK).count(),
        "reserved":    (_active_v.aggregate(r=Sum("reserved_stock"))["r"] or 0),
        "total_stock": (_active_v.aggregate(s=Sum("stock"))["s"] or 0),
    }

    stat_cards = [
        {"label": "Total Revenue", "value": revenue, "kind": "money", "icon": "rupee",
         "color": "#16a34a", "bg": "#f0fdf4", "growth": "+12.4% this month", "up": True},
        {"label": "Total Orders", "value": orders, "kind": "num", "icon": "bag",
         "color": "#ea580c", "bg": "#fff7ed", "growth": "+8.1% this month", "up": True},
        {"label": "Customers", "value": users, "kind": "num", "icon": "users",
         "color": "#0ea5e9", "bg": "#f0f9ff", "growth": "+18.2% this month", "up": True},
        {"label": "Products", "value": products, "kind": "num", "icon": "package",
         "color": "#16a34a", "bg": "#f0fdf4", "growth": "10 added this month", "up": True},
        {"label": "Conversion Rate", "value": "5.6%", "kind": "raw", "icon": "percent",
         "color": "#6b7280", "bg": "#f3f4f6", "growth": "+0.4% this month", "up": True},
        {"label": "Avg Order Value", "value": round(avg_order), "kind": "money", "icon": "trend",
         "color": "#6b7280", "bg": "#f3f4f6", "growth": "-2.1% this month", "up": False},
    ]

    pipeline_stages = [
        {"label": "Pending", "count": pipeline["pending"], "icon": "clock", "color": "#f59e0b"},
        {"label": "Confirmed", "count": pipeline["confirmed"], "icon": "check", "color": "#22c55e"},
        {"label": "Packed", "count": pipeline["packed"], "icon": "package", "color": "#3b82f6"},
        {"label": "Shipped", "count": pipeline["dispatched"], "icon": "truck", "color": "#8b5cf6"},
        {"label": "Delivered", "count": pipeline["delivered"], "icon": "check", "color": "#10b981"},
        {"label": "Returns", "count": pipeline["returns"], "icon": "refresh", "color": "#ef4444"},
    ]

    hour = timezone.localtime(now).hour
    greeting = "Good morning" if hour < 12 else "Good afternoon" if hour < 17 else "Good evening"

    return render(request, "panel/dashboard.html", _ctx(
        "/admin-panel",
        greeting=greeting,
        first_name=(request.user.first_name or "Admin"),
        today=timezone.localtime(now).strftime("%d %B %Y"),
        stat_cards=stat_cards, monthly=monthly,
        total_revenue=revenue, avg_monthly=avg_monthly, best_month=best_month,
        recent_orders=recent_orders, pipeline_stages=pipeline_stages,
        inv=inv, health_pct=health_pct, low_share=low_share, out_share=out_share,
        low_stock_items=low_stock_items, top_products=top_products,
        inv_analytics=inv_analytics,
    ))


# --------------------------------------------------------------------------- #
# Catalog management
# --------------------------------------------------------------------------- #
@admin_required
def products_list(request):
    q = request.GET.get("q", "").strip()
    cat_filter = request.GET.get("cat", "").strip()
    status_filter = request.GET.get("status", "").strip()
    sort = request.GET.get("sort", "position").strip()
    page_num = request.GET.get("page", 1)
    show_archived = request.GET.get("show", "") == "archived"

    back_url = "/admin-panel/products?show=archived" if show_archived else "/admin-panel/products"

    # ── Bulk POST actions ────────────────────────────────────────────
    if request.method == "POST":
        action = request.POST.get("bulk_action", "")
        ids = request.POST.getlist("product_ids")
        if ids and action:
            if show_archived:
                # Archived mode: IDs are Product PKs
                if action == "restore":
                    Product.objects.filter(pk__in=ids).update(status=Status.ACTIVE)
                    messages.success(request, f"Restored {len(ids)} product(s).")
                elif action == "delete":
                    Product.objects.filter(pk__in=ids).delete()
                    messages.success(request, f"Permanently deleted {len(ids)} product(s).")
            else:
                # Active mode: IDs are ProductVariant PKs
                if action == "delete":
                    ProductVariant.objects.filter(pk__in=ids).delete()
                    messages.success(request, f"Deleted {len(ids)} SKU(s).")
                elif action == "archive":
                    # Archive the parent product so it shows up in the Archived tab
                    product_ids = (ProductVariant.objects.filter(pk__in=ids)
                                   .values_list("product_id", flat=True).distinct())
                    Product.objects.filter(pk__in=product_ids).update(status=Status.INACTIVE)
                    messages.success(request, f"Archived {len(ids)} SKU(s).")
                elif action == "duplicate":
                    for vid in ids:
                        try:
                            v = ProductVariant.objects.get(pk=vid)
                            v.pk = None
                            v.va_code = gen_code("SKU")
                            v.variant = f"{v.variant} (Copy)"
                            v.reserved_stock = 0
                            v.save()
                        except ProductVariant.DoesNotExist:
                            pass
                    messages.success(request, f"Duplicated {len(ids)} SKU(s).")
        return redirect(back_url)

    # ── Stat cards: mutually exclusive, sum to total, no archived ───
    stat_base = ProductVariant.objects.filter(
        product__status=Status.ACTIVE, status=Status.ACTIVE
    )
    total_products = stat_base.count()
    in_stock  = stat_base.filter(stock_status=StockStatus.IN_STOCK).count()
    low_stock = stat_base.filter(stock_status=StockStatus.LOW_STOCK).count()
    out_stock = stat_base.filter(stock_status=StockStatus.OUT_OF_STOCK).count()
    cat_count  = Category.objects.count()
    avg_rating = round(
        Review.objects.filter(status="APPROVED").aggregate(a=Avg("rating"))["a"] or 0, 1
    )
    archived_count = Product.objects.filter(status=Status.INACTIVE).count()

    # ── Queryset ─────────────────────────────────────────────────────
    if show_archived:
        # Product-level view for archived tab
        qs = (Product.objects
              .filter(status=Status.INACTIVE)
              .select_related("category")
              .prefetch_related("variants", "reviews")
              .order_by("-updated_at"))
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    else:
        # Variant-as-product: one row per active variant of active products
        qs = (stat_base
              .select_related("product", "product__category")
              .prefetch_related("product__reviews"))
        if q:
            qs = qs.filter(
                Q(product__name__icontains=q) | Q(va_code__icontains=q) | Q(variant__icontains=q)
            )
        if cat_filter:
            qs = qs.filter(product__category__slug=cat_filter)
        if status_filter:
            qs = qs.filter(stock_status=status_filter)
        sort_map = {
            "name":      "product__name",
            "date_desc": "-product__created_at",
            "date_asc":  "product__created_at",
        }
        qs = qs.order_by(sort_map.get(sort, "product__position"))

    filtered_total = qs.count()
    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(page_num)

    return render(request, "panel/products.html", _ctx(
        "/admin-panel/products",
        products=page_obj, page_obj=page_obj,
        q=q, cat_filter=cat_filter, status_filter=status_filter, sort=sort,
        total=total_products, filtered_total=filtered_total,
        in_stock=in_stock, low_stock=low_stock, out_stock=out_stock,
        cat_count=cat_count, avg_rating=avg_rating,
        categories=Category.objects.order_by("position"),
        stock_statuses=StockStatus.choices,
        show_archived=show_archived,
        archived_count=archived_count,
    ))


@admin_required
def product_clone(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/products")
    orig = get_object_or_404(Product, pk=pk)
    orig.pk = None
    orig.code = gen_code("PRD")
    orig.name = f"{orig.name} (Copy)"
    orig.slug = slugify(orig.name)
    orig.position = next_position(Product)
    orig.save()
    messages.success(request, f"Cloned successfully.")
    return redirect("/admin-panel/products")


@admin_required
def product_export(request):
    q = request.GET.get("q", "").strip()
    cat_filter = request.GET.get("cat", "").strip()
    status_filter = request.GET.get("status", "").strip()
    ids = request.GET.getlist("ids")

    qs = (Product.objects.filter(status=Status.ACTIVE)
          .select_related("category")
          .prefetch_related("variants", "reviews").order_by("position"))
    if ids:
        qs = qs.filter(pk__in=ids)
    else:
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
        if cat_filter:
            qs = qs.filter(category__slug=cat_filter)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="products_export.csv"'
    response.write("﻿")  # BOM for Excel UTF-8

    writer = csv.writer(response)
    writer.writerow(["Code (SKU)", "Name", "Category", "MRP (₹)", "Price (₹)",
                     "Stock", "Stock Status", "Product Status", "Avg Rating", "Created"])
    for p in qs:
        v = p.default_variant
        total_stock = sum(pv.stock for pv in p.variants.all())
        writer.writerow([
            p.code, p.name, p.category.name,
            f"{v.mrp_price:.2f}" if v else "",
            f"{v.selling_price:.2f}" if v else "",
            total_stock,
            v.get_stock_status_display() if v else "",
            p.get_status_display(),
            p.avg_rating,
            p.created_at.strftime("%d %b %Y"),
        ])
    return response


@admin_required
def product_edit(request, pk=None):
    product = get_object_or_404(Product, pk=pk) if pk else None
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        data = dict(
            name=name,
            slug=request.POST.get("slug") or slugify(name),
            description=request.POST.get("description", ""),
            image=request.POST.get("image", "") or "/seed/placeholder.jpg",
            badge=request.POST.get("badge") or None,
            youtube_link=request.POST.get("youtube_link") or None,
            is_featured=bool(request.POST.get("is_featured")),
            top_seller=bool(request.POST.get("top_seller")),
            status=request.POST.get("status", Status.ACTIVE),
            position=int(request.POST.get("position") or 1),
            category=get_object_or_404(Category, pk=request.POST.get("category")),
        )
        pos = data["position"]
        try:
            if product:
                old_pos = product.position
                for k, v in data.items():
                    setattr(product, k, v)
                product.save()
                if pos != old_pos:
                    move_to_position(Product, product.pk, pos, old_pos)
                messages.success(request, "Product updated.")
            else:
                data["code"] = gen_code("PRD")
                with transaction.atomic():
                    if Product.objects.filter(position=pos).exists():
                        insert_at_position(Product, pos)
                    product = Product.objects.create(**data)
                messages.success(request, "Product created.")
        except IntegrityError:
            slug = data.get("slug", "")
            messages.error(request, f'The slug "{slug}" is already in use by another product. Please choose a different name or edit the slug manually.')
            return render(request, "panel/product_form.html", _ctx(
                "/admin-panel/products", product=product,
                categories=Category.objects.order_by("position"),
                statuses=Status.choices,
                form_data=request.POST,
                next_pos=next_position(Product) if not product else None,
            ))
        return redirect("/admin-panel/products")
    return render(request, "panel/product_form.html", _ctx(
        "/admin-panel/products", product=product,
        categories=Category.objects.order_by("position"),
        statuses=Status.choices,
        next_pos=next_position(Product) if not product else None))


@admin_required
def product_delete(request, pk):
    Product.objects.filter(pk=pk).delete()
    repack_positions(Product)
    messages.success(request, "Product deleted.")
    return redirect("/admin-panel/products")


@admin_required
def categories_list(request):
    q = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()
    sort = request.GET.get("sort", "position").strip()

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        pk = request.POST.get("pk", "").strip()
        data = dict(
            name=name,
            slug=request.POST.get("slug") or slugify(name),
            description=request.POST.get("description", "").strip(),
            image=request.POST.get("image") or None,
            position=int(request.POST.get("position") or 1),
            status=request.POST.get("status", Status.ACTIVE),
        )
        pos = data["position"]
        if pk:
            cat = get_object_or_404(Category, pk=pk)
            old_pos = cat.position
            for k, v in data.items():
                setattr(cat, k, v)
            cat.save()
            if pos != old_pos:
                move_to_position(Category, cat.pk, pos, old_pos)
            messages.success(request, f'Category "{name}" updated.')
        else:
            data["code"] = request.POST.get("code") or gen_code("CAT")
            with transaction.atomic():
                if Category.objects.filter(position=pos).exists():
                    insert_at_position(Category, pos)
                Category.objects.create(**data)
            messages.success(request, f'Category "{name}" added.')
        return redirect("/admin-panel/categories")

    all_cats = Category.objects.annotate(n=Count("products"))
    total_cats = all_cats.count()
    active_count = all_cats.filter(status=Status.ACTIVE).count()
    inactive_count = all_cats.filter(status=Status.INACTIVE).count()
    total_products = Product.objects.filter(status=Status.ACTIVE).count()

    cats = all_cats
    if q:
        cats = cats.filter(Q(name__icontains=q) | Q(slug__icontains=q) | Q(code__icontains=q))
    if status_filter:
        cats = cats.filter(status=status_filter)
    sort_map = {
        "name":     "name",
        "products": "-n",
        "date":     "-created_at",
        "position": "position",
    }
    cats = cats.order_by(sort_map.get(sort, "position"))

    return render(request, "panel/categories.html", _ctx(
        "/admin-panel/categories",
        categories=cats, statuses=Status.choices,
        active_count=active_count, inactive_count=inactive_count,
        total_products=total_products, total_cats=total_cats,
        q=q, status_filter=status_filter, sort=sort,
        next_pos=next_position(Category),
    ))


@admin_required
def category_delete(request, pk):
    Category.objects.filter(pk=pk).delete()
    repack_positions(Category)
    messages.success(request, "Category deleted.")
    return redirect("/admin-panel/categories")


@admin_required
def category_export(request):
    q = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()
    cats = Category.objects.annotate(n=Count("products"))
    if q:
        cats = cats.filter(Q(name__icontains=q) | Q(slug__icontains=q) | Q(code__icontains=q))
    if status_filter:
        cats = cats.filter(status=status_filter)
    cats = cats.order_by("position")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="categories_export.csv"'
    response.write("﻿")
    writer = csv.writer(response)
    writer.writerow(["Code", "Name", "Description", "Slug", "Position", "Status", "Products", "Created"])
    for c in cats:
        writer.writerow([
            c.code, c.name, c.description, c.slug, c.position,
            c.get_status_display(), c.n, c.created_at.strftime("%d %b %Y"),
        ])
    return response


@admin_required
def variants_list(request):
    if request.method == "POST":
        product = get_object_or_404(Product, pk=request.POST.get("product"))
        raw_pos = request.POST.get("position", "").strip()
        pos = int(raw_pos) if raw_pos.isdigit() else next_position(ProductVariant, product=product)
        with transaction.atomic():
            if ProductVariant.objects.filter(product=product, position=pos).exists():
                insert_at_position(ProductVariant, pos, product=product)
            ProductVariant.objects.create(
                product=product, va_code=request.POST.get("va_code") or gen_code("VA"),
                variant=request.POST.get("variant", ""),
                short_name=request.POST.get("short_name") or request.POST.get("variant", ""),
                selling_price=float(request.POST.get("selling_price") or 0),
                mrp_price=float(request.POST.get("mrp_price") or 0),
                stock=int(request.POST.get("stock") or 0),
                position=pos)
        messages.success(request, "Variant added.")
        return redirect("/admin-panel/variants")

    q              = request.GET.get("q", "").strip()
    product_filter = request.GET.get("product", "").strip()
    status_filter  = request.GET.get("status", "").strip()
    page_num       = request.GET.get("page", 1)

    all_variants = (ProductVariant.objects
                    .select_related("product", "product__category")
                    .order_by("product__name", "position"))

    total_v     = all_variants.count()
    in_stock_v  = all_variants.filter(stock_status=StockStatus.IN_STOCK).count()
    low_stock_v = all_variants.filter(stock_status=StockStatus.LOW_STOCK).count()
    out_stock_v = all_variants.filter(stock_status=StockStatus.OUT_OF_STOCK).count()

    if q:
        all_variants = all_variants.filter(
            Q(variant__icontains=q) | Q(va_code__icontains=q) | Q(product__name__icontains=q)
        )
    if product_filter:
        all_variants = all_variants.filter(product_id=product_filter)
    if status_filter:
        all_variants = all_variants.filter(stock_status=status_filter)

    filtered_total = all_variants.count()
    paginator = Paginator(all_variants, 10)
    page_obj  = paginator.get_page(page_num)

    flat_rows = []
    prev_pid  = None
    for v in page_obj:
        flat_rows.append({"variant": v, "is_first": v.product_id != prev_pid})
        prev_pid = v.product_id

    return render(request, "panel/variants.html", _ctx(
        "/admin-panel/variants",
        flat_rows=flat_rows, page_obj=page_obj,
        products=Product.objects.order_by("name"),
        total_v=total_v, in_stock_v=in_stock_v,
        low_stock_v=low_stock_v, out_stock_v=out_stock_v,
        filtered_total=filtered_total,
        q=q, product_filter=product_filter, status_filter=status_filter,
        stock_statuses=StockStatus.choices,
    ))


@admin_required
def variant_edit(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/variants")
    v = get_object_or_404(ProductVariant, pk=pk)
    old_pos = v.position
    v.variant       = request.POST.get("variant", v.variant).strip() or v.variant
    v.selling_price = float(request.POST.get("selling_price") or v.selling_price)
    v.mrp_price     = float(request.POST.get("mrp_price") or v.mrp_price)
    v.stock         = int(request.POST.get("stock") if request.POST.get("stock") != "" else v.stock)
    new_pos         = int(request.POST.get("position") or old_pos)
    v.position      = new_pos
    v.save()
    if new_pos != old_pos:
        move_to_position(ProductVariant, v.pk, new_pos, old_pos, product=v.product)
    messages.success(request, f'Variant "{v.variant}" updated.')
    return redirect("/admin-panel/variants")


@admin_required
def variant_next_position(request):
    product_id = request.GET.get("product", "").strip()
    if product_id:
        pos = next_position(ProductVariant, product_id=product_id)
    else:
        pos = next_position(ProductVariant)
    return _JsonResponse({"next_pos": pos})


@admin_required
def variant_delete(request, pk):
    v = get_object_or_404(ProductVariant, pk=pk)
    product_id = v.product_id
    redirect_url = "/admin-panel/products" if request.GET.get("from") == "products" else "/admin-panel/variants"
    v.delete()
    repack_positions(ProductVariant, product_id=product_id)
    messages.success(request, "Variant deleted.")
    return redirect(redirect_url)


@admin_required
def variant_clone(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/products")
    v = get_object_or_404(ProductVariant, pk=pk)
    product_id = v.product_id
    v.pk = None
    v.va_code = gen_code("SKU")
    v.variant = f"{v.variant} (Copy)"
    v.reserved_stock = 0
    v.position = next_position(ProductVariant, product_id=product_id)
    v.save()
    messages.success(request, f'Variant cloned as "{v.variant}".')
    return redirect("/admin-panel/products")


@admin_required
def inventory(request):
    if request.method == "POST":
        action = request.POST.get("action", "save_stock")

        # ── Save / update global threshold ──────────────────────────────────
        if action == "save_threshold":
            in_min = max(int(request.POST.get("in_stock_min") or 51), 1)
            low_min = max(int(request.POST.get("low_stock_min") or 1), 0)
            if low_min >= in_min:
                messages.error(request, "Low Stock threshold must be less than In Stock threshold.")
                return redirect("/admin-panel/inventory")
            try:
                t = StockThreshold.objects.get(product__isnull=True)
                t.in_stock_min = in_min
                t.low_stock_min = low_min
                t.save()
            except StockThreshold.DoesNotExist:
                StockThreshold.objects.create(product=None, in_stock_min=in_min, low_stock_min=low_min)
            messages.success(request, f"Threshold updated: ≥{in_min} In Stock · ≥{low_min} Low Stock.")
            return redirect("/admin-panel/inventory")

        # ── Apply threshold rules to all variants in bulk ────────────────────
        if action == "apply_thresholds":
            updated = 0
            for v in ProductVariant.objects.select_related("product").all():
                threshold = StockThreshold.get_for_product(v.product)
                new_status = threshold.compute_status(v.stock)
                if v.stock_status != new_status:
                    old_status = v.stock_status
                    v.stock_status = new_status
                    v.save(update_fields=["stock_status"])
                    StockStatusHistory.objects.create(
                        variant=v,
                        old_status=old_status,
                        new_status=new_status,
                        old_stock=v.stock,
                        new_stock=v.stock,
                        note="Bulk threshold apply",
                    )
                    updated += 1
            messages.success(request, f"Rules applied — {updated} variant{'s' if updated != 1 else ''} updated.")
            return redirect("/admin-panel/inventory")

        # ── Save individual variant stock (auto-compute status) ──────────────
        v = get_object_or_404(ProductVariant, pk=request.POST.get("variant_id"))
        old_stock = v.stock
        old_status = v.stock_status
        new_stock = max(int(request.POST.get("stock") or 0), 0)

        threshold = StockThreshold.get_for_product(v.product)
        new_status = threshold.compute_status(new_stock)

        v.stock = new_stock
        v.stock_status = new_status
        v.save()

        if old_stock != new_stock or old_status != new_status:
            StockStatusHistory.objects.create(
                variant=v,
                old_status=old_status,
                new_status=new_status,
                old_stock=old_stock,
                new_stock=new_stock,
            )

        messages.success(
            request,
            f"{v.product.name} ({v.variant}): {new_stock} units → {v.get_stock_status_display()}",
        )
        return redirect("/admin-panel/inventory")

    # ── GET ──────────────────────────────────────────────────────────────────
    now = timezone.now()
    thirty_days_ago = now - timezone.timedelta(days=30)
    variants = (ProductVariant.objects
                .select_related("product", "product__category")
                .order_by("stock_status", "stock"))
    total_skus = variants.count()
    in_stock_c = variants.filter(stock_status=StockStatus.IN_STOCK).count()
    low_stock_c = variants.filter(stock_status=StockStatus.LOW_STOCK).count()
    out_stock_c = variants.filter(stock_status=StockStatus.OUT_OF_STOCK).count()

    sold_map = {
        r["variant_id"]: r["sold"]
        for r in (OrderItem.objects
                  .filter(order__created_at__gte=thirty_days_ago)
                  .values("variant_id")
                  .annotate(sold=Sum("qty")))
    }

    # Last history entry per variant — cross-database via Max(id) trick
    from django.db.models import Max as _Max
    latest_ids = list(
        StockStatusHistory.objects.values("variant_id").annotate(mid=_Max("id")).values_list("mid", flat=True)
    )
    last_history_map = {
        h.variant_id: h
        for h in StockStatusHistory.objects.filter(id__in=latest_ids)
    }

    variants_data = []
    max_stock = max((v.stock for v in variants), default=1) or 1
    for v in variants:
        variants_data.append({
            "variant": v,
            "sold_30d": sold_map.get(v.id, 0),
            "stock_pct": min(round(v.stock / max_stock * 100), 100),
            "last_history": last_history_map.get(v.id),
        })

    health_pct = round(in_stock_c / total_skus * 100) if total_skus else 0
    low_share = round(low_stock_c / total_skus * 100) if total_skus else 0
    out_share = round(out_stock_c / total_skus * 100) if total_skus else 0
    alert_items = [r for r in variants_data if r["variant"].stock_status in (
        StockStatus.LOW_STOCK, StockStatus.OUT_OF_STOCK)][:6]

    # Global threshold (used for JS preview too)
    try:
        global_threshold = StockThreshold.objects.get(product__isnull=True)
    except StockThreshold.DoesNotExist:
        global_threshold = StockThreshold(in_stock_min=51, low_stock_min=1)

    # Recent history (last 10 changes, all variants)
    recent_history = StockStatusHistory.objects.select_related("variant__product").order_by("-changed_at")[:10]

    total_reserved = (variants.aggregate(r=Sum("reserved_stock"))["r"] or 0)

    return render(request, "panel/inventory.html", _ctx(
        "/admin-panel/inventory",
        variants_data=variants_data,
        total_skus=total_skus, in_stock_c=in_stock_c, low_stock_c=low_stock_c,
        out_stock_c=out_stock_c, total_reserved=total_reserved,
        health_pct=health_pct, low_share=low_share,
        out_share=out_share, alert_items=alert_items,
        global_threshold=global_threshold,
        recent_history=recent_history,
    ))


# --------------------------------------------------------------------------- #
# Orders management
# --------------------------------------------------------------------------- #
TAB_STATUS_MAP = {
    "pending": [OrderStatus.PROCESSING],
    "processing": [OrderStatus.ORDER_CONFIRMED, OrderStatus.PACKED],
    "shipped": [OrderStatus.DISPATCHED],
    "delivered": [OrderStatus.DELIVERED],
    "cancelled": [OrderStatus.CANCELLED],
    "returns": [OrderStatus.REFUNDED],
}

TAB_LABELS = {
    "": "All",
    "pending": "Pending",
    "processing": "Processing",
    "shipped": "Shipped",
    "delivered": "Delivered",
    "cancelled": "Cancelled",
    "returns": "Returns",
}


@admin_required
def orders_list(request):
    from django.db.models import Q
    tab = request.GET.get("tab", "")
    q = request.GET.get("q", "").strip()
    channel = request.GET.get("channel", "")

    qs = (Order.objects.select_related("user")
          .prefetch_related("items__product").order_by("-created_at"))

    if tab and tab in TAB_STATUS_MAP:
        qs = qs.filter(status__in=TAB_STATUS_MAP[tab])
    elif request.GET.get("status"):
        qs = qs.filter(status=request.GET.get("status"))

    if q:
        qs = qs.filter(Q(order_id__icontains=q) | Q(user__first_name__icontains=q) |
                       Q(user__last_name__icontains=q) | Q(user__email__icontains=q))
    if channel:
        qs = qs.filter(channel=channel)

    # Tab counts
    tab_counts = {
        "": Order.objects.count(),
        "pending": Order.objects.filter(status__in=TAB_STATUS_MAP["pending"]).count(),
        "processing": Order.objects.filter(status__in=TAB_STATUS_MAP["processing"]).count(),
        "shipped": Order.objects.filter(status__in=TAB_STATUS_MAP["shipped"]).count(),
        "delivered": Order.objects.filter(status__in=TAB_STATUS_MAP["delivered"]).count(),
        "cancelled": Order.objects.filter(status__in=TAB_STATUS_MAP["cancelled"]).count(),
        "returns": Order.objects.filter(status__in=TAB_STATUS_MAP["returns"]).count(),
    }
    # Stat cards for header
    stat_cards = [
        {"label": "Total Orders", "value": tab_counts[""], "sub": "+8.1% this month", "up": True, "icon": "bag", "color": "#ea580c", "bg": "#fff7ed"},
        {"label": "Pending", "value": tab_counts["pending"], "sub": "Awaiting confirmation", "up": None, "icon": "clock", "color": "#f59e0b", "bg": "#fffbeb"},
        {"label": "Processing", "value": tab_counts["processing"], "sub": "Being packed", "up": None, "icon": "loader", "color": "#3b82f6", "bg": "#eff6ff"},
        {"label": "Shipped", "value": tab_counts["shipped"], "sub": "Out for delivery", "up": None, "icon": "truck", "color": "#8b5cf6", "bg": "#f5f3ff"},
        {"label": "Delivered", "value": tab_counts["delivered"], "sub": f"{round(tab_counts['delivered']/tab_counts[''] * 100, 1) if tab_counts[''] else 0}% of all orders", "up": True, "icon": "check-circle", "color": "#16a34a", "bg": "#f0fdf4"},
        {"label": "Cancelled / Returns", "value": tab_counts["cancelled"] + tab_counts["returns"], "sub": f"{round((tab_counts['cancelled']+tab_counts['returns'])/tab_counts[''] * 100, 1) if tab_counts[''] else 0}% rate", "up": False, "icon": "x-circle", "color": "#dc2626", "bg": "#fef2f2"},
    ]
    import json
    return render(request, "panel/orders.html", _ctx(
        "/admin-panel/orders", orders=qs, tab=tab, q=q, channel=channel,
        tab_counts=tab_counts, stat_cards=stat_cards, tab_labels=TAB_LABELS,
        channels=OrderChannel.choices,
        order_statuses=OrderStatus.choices, payment_statuses=PaymentStatus.choices,
        workflow_json=json.dumps(WORKFLOW_TRANSITIONS),
        cancellable_statuses=list(WORKFLOW_TRANSITIONS.keys())[:-2],  # all except CANCELLED/REFUNDED
        refundable_statuses=[OrderStatus.DELIVERED, OrderStatus.CANCELLED],
    ))


def _log_event(order, title, description="", actor=None):
    OrderEvent.objects.create(
        order=order, title=title, description=description,
        actor_name=actor.full_name if actor else "System",
    )


def _notify_customer(order, title, message):
    Notification.objects.create(user=order.user, title=title, message=message)


@admin_required
def order_detail(request, pk):
    order = get_object_or_404(
        Order.objects.select_related("user", "coupon")
             .prefetch_related("items__product", "items__variant", "items__combo",
                               "notes", "events", "refunds"),
        pk=pk,
    )
    items = list(order.items.all())

    # Group combo items by combo_id for display
    combos = {}
    regular_items = []
    for it in items:
        if it.combo_id:
            combos.setdefault(it.combo_id, {"combo": it.combo, "items": []})["items"].append(it)
        else:
            regular_items.append(it)

    import json as _json
    try:
        addr = _json.loads(order.shipping_address)
    except Exception:
        addr = {"address": order.shipping_address}

    status_labels = dict(OrderStatus.choices)
    next_statuses = [
        (s, status_labels.get(s, s))
        for s in WORKFLOW_TRANSITIONS.get(order.status, [])
    ]
    user = order.user
    user_orders_count = Order.objects.filter(user=user).count()
    user_total_spend = Order.objects.filter(
        user=user, payment_status=PaymentStatus.PAID
    ).aggregate(total=Sum("grand_total"))["total"] or 0

    return render(request, "panel/order_detail.html", _ctx(
        "/admin-panel/orders",
        order=order,
        items=items,
        combos=combos,
        regular_items=regular_items,
        addr=addr,
        notes=list(order.notes.all()),
        events=list(order.events.all()),
        refunds=list(order.refunds.all()),
        next_statuses=next_statuses,
        order_statuses=OrderStatus.choices,
        payment_statuses=PaymentStatus.choices,
        refundable=order.status in (OrderStatus.DELIVERED, OrderStatus.CANCELLED),
        cancellable=_valid_transition(order.status, OrderStatus.CANCELLED),
        user_orders_count=user_orders_count,
        user_total_spend=user_total_spend,
        PaymentMode=PaymentMode,
        OrderStatus=OrderStatus,
        PaymentStatus=PaymentStatus,
    ))


@admin_required
def order_note_add(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method != "POST":
        return redirect(f"/admin-panel/orders/{pk}")
    text = request.POST.get("text", "").strip()
    if text:
        is_internal = request.POST.get("is_internal", "1") == "1"
        OrderNote.objects.create(
            order=order, text=text, is_internal=is_internal,
            created_by_name=request.user.full_name,
        )
        _log_event(order, "Note Added", text[:80], actor=request.user)
        messages.success(request, "Note added.")
    return redirect(f"/admin-panel/orders/{pk}")


@admin_required
def order_update(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method != "POST":
        return redirect(f"/admin-panel/orders/{pk}")

    new_status = request.POST.get("status", order.status)
    new_payment = request.POST.get("payment_status", order.payment_status)
    prev_status = order.status

    # Validate workflow transition (allow no-op)
    if new_status != order.status and not _valid_transition(order.status, new_status):
        messages.error(
            request,
            f"Invalid transition: {order.get_status_display()} → {new_status}. "
            f"Follow the fixed workflow.",
        )
        return redirect(f"/admin-panel/orders/{pk}")

    # Stock side-effects
    if new_status != order.status:
        if new_status == OrderStatus.ORDER_CONFIRMED:
            _deduct_stock_for_order(order)
        elif new_status == OrderStatus.DELIVERED:
            _release_reserved_on_deliver(order)

    order.status = new_status
    order.payment_status = new_payment
    order.save()

    if new_status != prev_status:
        display = dict(OrderStatus.choices).get(new_status, new_status)
        _log_event(order, f"Status → {display}", actor=request.user)
        _notify_customer(
            order,
            f"Order {order.order_id} Update",
            f"Your order {order.order_id} status has been updated to: {display}.",
        )

    messages.success(request, f"Order {order.order_id} → {order.get_status_display()}.")
    return redirect(f"/admin-panel/orders/{pk}")


@admin_required
def order_cancel(request, pk):
    """Cancel an order and restore stock if it was already deducted."""
    order = get_object_or_404(Order, pk=pk)
    if request.method != "POST":
        return redirect(f"/admin-panel/orders/{pk}")

    if not _valid_transition(order.status, OrderStatus.CANCELLED):
        messages.error(request, f"Order {order.order_id} cannot be cancelled at this stage.")
        return redirect(f"/admin-panel/orders/{pk}")

    _restore_stock_for_order(order)
    order.status = OrderStatus.CANCELLED
    order.save()

    _log_event(order, "Order Cancelled", "Stock restored to inventory.", actor=request.user)
    _notify_customer(
        order,
        f"Order {order.order_id} Cancelled",
        f"Your order {order.order_id} has been cancelled. If you paid online, a refund will be initiated.",
    )
    messages.success(request, f"Order {order.order_id} cancelled. Stock restored.")
    return redirect(f"/admin-panel/orders/{pk}")


@admin_required
def order_refund(request, pk):
    """Issue a full or partial refund for an order."""
    order = get_object_or_404(Order, pk=pk)
    if request.method != "POST":
        return redirect(f"/admin-panel/orders/{pk}")

    refund_type = request.POST.get("refund_type", OrderRefund.FULL)
    try:
        amount = float(request.POST.get("amount") or order.grand_total)
    except ValueError:
        amount = order.grand_total
    amount = min(max(amount, 0), order.grand_total)
    reason = request.POST.get("reason", "").strip()

    OrderRefund.objects.create(order=order, refund_type=refund_type, amount=amount, reason=reason)
    order.payment_status = PaymentStatus.REFUNDED
    if refund_type == OrderRefund.FULL:
        order.status = OrderStatus.REFUNDED
    order.save()

    from core.utils import inr
    _log_event(order, f"Refund Issued — ₹{amount:.0f} ({refund_type})", reason, actor=request.user)
    _notify_customer(
        order,
        f"Refund for Order {order.order_id}",
        f"A refund of ₹{amount:.0f} has been initiated for order {order.order_id}. "
        f"It will reflect in your account within 5–7 business days.",
    )
    messages.success(request, f"Refund of {inr(amount)} ({refund_type}) recorded for {order.order_id}.")
    return redirect(f"/admin-panel/orders/{pk}")


# --------------------------------------------------------------------------- #
# Customer classification helpers
# --------------------------------------------------------------------------- #
# Thresholds — change these values to adjust automatic tagging logic.
CUST_VIP_SPEND   = 15000   # lifetime spend ≥ this → VIP
CUST_VIP_ORDERS  = 10      # order count ≥ this → VIP
CUST_REPEAT_MIN  = 3       # order count ≥ this (and not VIP) → Repeat
# orders < CUST_REPEAT_MIN and not VIP → New

AVATAR_COLORS = [
    "bg-brand-100 text-brand-700",
    "bg-blue-100 text-blue-700",
    "bg-purple-100 text-purple-700",
    "bg-orange-100 text-orange-700",
    "bg-teal-100 text-teal-700",
    "bg-rose-100 text-rose-700",
    "bg-indigo-100 text-indigo-700",
    "bg-amber-100 text-amber-700",
]


def _classify(order_count, total_spend, manual_tags_str):
    """Return list of tag strings for a customer."""
    oc    = order_count or 0
    spend = total_spend or 0
    tags  = []
    if spend >= CUST_VIP_SPEND or oc >= CUST_VIP_ORDERS:
        tags.append("VIP")
    if oc >= CUST_REPEAT_MIN:
        tags.append("Repeat")
    elif oc > 0:
        tags.append("New")
    # Manual tags (dedupe, preserve order)
    for t in (manual_tags_str or "").split(","):
        t = t.strip()
        if t and t not in tags:
            tags.append(t)
    return tags


def _decorate_users(page_iterable):
    """Attach avatar_style, initials, tags to each user object in-place."""
    for u in page_iterable:
        u.avatar_style = AVATAR_COLORS[u.id % len(AVATAR_COLORS)]
        u.initials = (
            ((u.first_name or "")[:1] + (u.last_name or "")[:1]).upper()
            or (u.email or "?")[:1].upper()
        )
        u.tags = _classify(
            getattr(u, "order_count", 0),
            getattr(u, "total_spend", 0),
            u.manual_tags,
        )
    return page_iterable


# --------------------------------------------------------------------------- #
# Customer List
# --------------------------------------------------------------------------- #
@admin_required
def users_list(request):
    # ── Bulk POST actions ──────────────────────────────────────────────────
    if request.method == "POST":
        action = request.POST.get("action", "")
        ids    = request.POST.getlist("ids")
        if ids:
            if action == "activate":
                User.objects.filter(pk__in=ids).update(status=Status.ACTIVE)
            elif action == "deactivate":
                User.objects.filter(pk__in=ids).update(status=Status.INACTIVE)
            elif action == "delete":
                User.objects.filter(pk__in=ids, user_type=UserType.USER).delete()
                messages.success(request, f"Deleted {len(ids)} customer(s).")
            elif action == "export_selected":
                return _customer_csv(User.objects.filter(pk__in=ids))
        return redirect("/admin-panel/users")

    # ── Filters ────────────────────────────────────────────────────────────
    q         = request.GET.get("q", "").strip()
    segment   = request.GET.get("segment", "")   # all | vip | repeat | new
    sort      = request.GET.get("sort", "orders")
    status_f  = request.GET.get("status", "")

    base_qs = (
        User.objects.filter(user_type=UserType.USER)
        .annotate(
            order_count    = Count("orders", distinct=True),
            total_spend    = Sum("orders__grand_total"),
            last_order_date= Max("orders__created_at"),
        )
    )

    if q:
        base_qs = base_qs.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(email__icontains=q)      | Q(mobile__icontains=q) |
            Q(city__icontains=q)
        )
    if status_f:
        base_qs = base_qs.filter(status=status_f)

    sort_map = {
        "orders":   "-order_count",
        "spend":    "-total_spend",
        "newest":   "-created_at",
        "oldest":   "created_at",
        "last":     "-last_order_date",
        "alpha":    "first_name",
    }
    base_qs = base_qs.order_by(sort_map.get(sort, "-order_count"))

    # Segment filter: need classification → evaluate to list, then paginate
    all_users  = list(base_qs)
    _decorate_users(all_users)

    if segment == "vip":
        all_users = [u for u in all_users if "VIP" in u.tags]
    elif segment == "repeat":
        all_users = [u for u in all_users if "Repeat" in u.tags and "VIP" not in u.tags]
    elif segment == "new":
        all_users = [u for u in all_users if "New" in u.tags]

    # ── Stats ──────────────────────────────────────────────────────────────
    now         = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    total_cust  = User.objects.filter(user_type=UserType.USER).count()
    new_month   = User.objects.filter(user_type=UserType.USER,
                                      created_at__gte=month_start).count()
    vip_count   = sum(1 for u in all_users if "VIP" in u.tags)
    spends      = [u.total_spend for u in all_users if u.total_spend]
    avg_clv     = round(sum(spends) / len(spends)) if spends else 0

    paginator = Paginator(all_users, 10)
    page_obj  = paginator.get_page(request.GET.get("page", 1))

    return render(request, "panel/users.html", {
        "active_path":   "/admin-panel/users",
        "users":         page_obj,
        "page_obj":      page_obj,
        "filtered_total": len(all_users),
        "q": q, "segment": segment, "sort": sort, "status_f": status_f,
        "total_cust": total_cust,
        "new_month":  new_month,
        "vip_count":  vip_count,
        "avg_clv":    avg_clv,
        "vip_spend":  CUST_VIP_SPEND,
        "vip_orders": CUST_VIP_ORDERS,
        "repeat_min": CUST_REPEAT_MIN,
    })


def _customer_csv(qs):
    """Return CSV HttpResponse for a queryset of User objects."""
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="customers.csv"'
    w = csv.writer(resp)
    w.writerow(["ID", "Name", "Email", "Mobile", "City", "Status",
                "Orders", "Total Spend", "Joined"])
    qs = qs.annotate(
        order_count=Count("orders", distinct=True),
        total_spend=Sum("orders__grand_total"),
    )
    for u in qs:
        w.writerow([
            u.pk, u.full_name, u.email, u.mobile, u.city or "",
            u.status, u.order_count or 0,
            round(u.total_spend or 0, 2),
            u.created_at.strftime("%Y-%m-%d"),
        ])
    return resp


@admin_required
def user_export(request):
    qs = User.objects.filter(user_type=UserType.USER)
    return _customer_csv(qs)


# --------------------------------------------------------------------------- #
# Customer Detail
# --------------------------------------------------------------------------- #
@admin_required
def user_detail(request, pk):
    customer = get_object_or_404(User, pk=pk, user_type=UserType.USER)

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "toggle_status":
            customer.status = (
                Status.INACTIVE if customer.status == Status.ACTIVE else Status.ACTIVE
            )
            customer.save(update_fields=["status"])
            messages.success(request, f"Account {'activated' if customer.status == Status.ACTIVE else 'deactivated'}.")
            return redirect(f"/admin-panel/users/{pk}")

        if action == "update_tags":
            customer.manual_tags = request.POST.get("manual_tags", "").strip()
            customer.save(update_fields=["manual_tags"])
            messages.success(request, "Tags updated.")
            return redirect(f"/admin-panel/users/{pk}")

        if action == "add_note":
            note_text = request.POST.get("note", "").strip()
            if note_text:
                CustomerNote.objects.create(
                    user=customer,
                    note=note_text,
                    author=request.user,
                )
                messages.success(request, "Note added.")
            return redirect(f"/admin-panel/users/{pk}#notes")

        if action == "delete_note":
            note_id = request.POST.get("note_id")
            CustomerNote.objects.filter(pk=note_id, user=customer).delete()
            return redirect(f"/admin-panel/users/{pk}#notes")

        if action == "update_info":
            customer.first_name = request.POST.get("first_name", "").strip() or None
            customer.last_name  = request.POST.get("last_name",  "").strip() or None
            customer.mobile     = request.POST.get("mobile", "").strip() or None
            customer.city       = request.POST.get("city", "").strip() or None
            dob_str = request.POST.get("date_of_birth", "").strip()
            if dob_str:
                try:
                    from datetime import date as _ddate
                    customer.date_of_birth = _ddate.fromisoformat(dob_str)
                except ValueError:
                    pass
            else:
                customer.date_of_birth = None
            gender = request.POST.get("gender", "").strip()
            customer.gender = gender or None
            try:
                customer.loyalty_points = int(request.POST.get("loyalty_points", customer.loyalty_points))
            except (ValueError, TypeError):
                pass
            try:
                customer.wallet_balance = float(request.POST.get("wallet_balance", customer.wallet_balance))
            except (ValueError, TypeError):
                pass
            customer.save(update_fields=[
                "first_name", "last_name", "mobile", "city",
                "date_of_birth", "gender", "loyalty_points", "wallet_balance", "updated_at",
            ])
            messages.success(request, "Customer info updated.")
            return redirect(f"/admin-panel/users/{pk}")

    # ── Aggregate data ─────────────────────────────────────────────────────
    orders_qs = (Order.objects.filter(user=customer)
                 .select_related("user")
                 .order_by("-created_at"))
    order_count  = orders_qs.count()
    agg          = orders_qs.aggregate(total=Sum("grand_total"), avg=Avg("grand_total"))
    total_spend  = round(agg["total"] or 0, 2)
    avg_order    = round(agg["avg"] or 0, 2)
    last_order   = orders_qs.first()
    reviews_qs   = Review.objects.filter(user=customer).select_related("product").order_by("-created_at")
    addresses_qs = customer.addresses.all()
    notes_qs     = customer.customer_notes.select_related("author").all()

    tags = _classify(order_count, total_spend, customer.manual_tags)
    customer.avatar_style = AVATAR_COLORS[customer.id % len(AVATAR_COLORS)]
    customer.initials = (
        ((customer.first_name or "")[:1] + (customer.last_name or "")[:1]).upper()
        or (customer.email or "?")[:1].upper()
    )

    paginator   = Paginator(orders_qs, 8)
    orders_page = paginator.get_page(request.GET.get("opage", 1))

    return render(request, "panel/user_detail.html", {
        "active_path":  "/admin-panel/users",
        "customer":     customer,
        "tags":         tags,
        "order_count":  order_count,
        "total_spend":  total_spend,
        "avg_order":    avg_order,
        "last_order":   last_order,
        "orders":       orders_page,
        "orders_page":  orders_page,
        "reviews":      reviews_qs[:10],
        "addresses":    addresses_qs,
        "notes":        notes_qs,
    })


@admin_required
def user_toggle_status(request, pk):
    from django.http import JsonResponse
    customer = get_object_or_404(User, pk=pk, user_type=UserType.USER)
    customer.status = Status.INACTIVE if customer.status == Status.ACTIVE else Status.ACTIVE
    customer.save(update_fields=["status"])
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"ok": True, "status": customer.status})
    return redirect("/admin-panel/users")


@admin_required
def loyalty_members(request):
    qs = (User.objects.filter(user_type=UserType.USER)
          .annotate(order_count=Count("orders"), total_spend=Sum("orders__grand_total"))
          .order_by("-total_spend"))
    return render(request, "panel/loyalty_members.html", _ctx(
        "/admin-panel/loyalty", members=qs))


# --------------------------------------------------------------------------- #
# Coming-soon helper
# --------------------------------------------------------------------------- #
def _wip(request, module_name, active_path=None):
    return render(request, "panel/coming_soon.html",
                  _ctx(active_path or "/admin-panel", module_name=module_name))


@admin_required
def newsletter_view(request):
    return _wip(request, "Newsletter", "/admin-panel/marketing/newsletter")


# --------------------------------------------------------------------------- #
# Analytics views
# --------------------------------------------------------------------------- #
@admin_required
def sales_reports(request):
    return _wip(request, "Sales Report", "/admin-panel/reports")


@admin_required
def revenue_view(request):
    return _wip(request, "Revenue Analytics", "/admin-panel/revenue")


@admin_required
def performance_view(request):
    return _wip(request, "Performance", "/admin-panel/performance")


@admin_required
def admin_settings(request):
    from store.models import SiteSettings, IntegrationConfig
    from accounts.models import User as _User, UserType as _UT

    tab = request.GET.get("tab", "store")

    if request.method == "POST":
        section = request.POST.get("section", "")
        cfg = SiteSettings.get()
        P = request.POST

        if section == "store_identity":
            logo = request.FILES.get("logo")
            if logo:
                from core.cloudinary_storage import upload_file as _cl_up
                url, err = _cl_up(logo.read(), logo.name, logo.content_type)
                if url:
                    cfg.logo_url = url
                else:
                    messages.warning(request, f"Logo upload failed: {err or 'Cloud storage is not configured.'}")
            cfg.store_name    = P.get("store_name", cfg.store_name).strip()
            cfg.store_tagline = P.get("store_tagline", cfg.store_tagline).strip()
            cfg.business_email = P.get("business_email", cfg.business_email).strip()
            cfg.support_phone  = P.get("support_phone", cfg.support_phone).strip()
            cfg.store_address  = P.get("store_address", cfg.store_address).strip()
            cfg.save()
            messages.success(request, "Store identity saved.")
            return redirect("/admin-panel/settings?tab=store")

        if section == "regional":
            cfg.currency        = P.get("currency", cfg.currency)
            cfg.timezone        = P.get("timezone", cfg.timezone)
            cfg.language        = P.get("language", cfg.language)
            cfg.date_format     = P.get("date_format", cfg.date_format)
            cfg.weight_unit     = P.get("weight_unit", cfg.weight_unit)
            cfg.order_id_prefix = P.get("order_id_prefix", cfg.order_id_prefix).strip()
            cfg.save()
            messages.success(request, "Regional settings saved.")
            return redirect("/admin-panel/settings?tab=store")

        if section == "shipping":
            try:
                cfg.free_shipping_above     = int(P.get("free_shipping_above", cfg.free_shipping_above))
                cfg.default_shipping_charge = int(P.get("default_shipping_charge", cfg.default_shipping_charge))
            except ValueError:
                pass
            cfg.processing_time        = P.get("processing_time", cfg.processing_time).strip()
            cfg.estimated_delivery     = P.get("estimated_delivery", cfg.estimated_delivery).strip()
            cfg.cod_enabled            = "cod_enabled" in P
            cfg.show_delivery_estimate = "show_delivery_estimate" in P
            cfg.international_shipping = "international_shipping" in P
            cfg.save()
            messages.success(request, "Shipping & delivery settings saved.")
            return redirect("/admin-panel/settings?tab=store")

        if section == "tax":
            cfg.gstin                   = P.get("gstin", cfg.gstin).strip()
            cfg.default_gst_rate        = P.get("default_gst_rate", cfg.default_gst_rate)
            cfg.prices_inclusive_of_gst = "prices_inclusive_of_gst" in P
            cfg.show_gst_in_invoice     = "show_gst_in_invoice" in P
            cfg.save()
            messages.success(request, "Tax configuration saved.")
            return redirect("/admin-panel/settings?tab=store")

        if section == "notif_store":
            cfg.notif_new_order       = "notif_new_order" in P
            cfg.notif_order_cancelled = "notif_order_cancelled" in P
            cfg.notif_refund_request  = "notif_refund_request" in P
            cfg.notif_order_delivered = "notif_order_delivered" in P
            cfg.notif_low_stock       = "notif_low_stock" in P
            cfg.notif_out_of_stock    = "notif_out_of_stock" in P
            cfg.notif_restock         = "notif_restock" in P
            cfg.notif_new_review      = "notif_new_review" in P
            cfg.save()
            messages.success(request, "Notification preferences saved.")
            return redirect("/admin-panel/settings?tab=store")

        if section == "notif_full":
            cfg.notif_new_order           = "notif_new_order" in P
            cfg.notif_order_cancelled     = "notif_order_cancelled" in P
            cfg.notif_refund_request      = "notif_refund_request" in P
            cfg.notif_order_delivered     = "notif_order_delivered" in P
            cfg.notif_low_stock           = "notif_low_stock" in P
            cfg.notif_out_of_stock        = "notif_out_of_stock" in P
            cfg.notif_restock             = "notif_restock" in P
            cfg.notif_new_review          = "notif_new_review" in P
            cfg.notif_customer_registered = "notif_customer_registered" in P
            cfg.notif_payment_success     = "notif_payment_success" in P
            cfg.notif_payment_failed      = "notif_payment_failed" in P
            cfg.notif_promotional         = "notif_promotional" in P
            cfg.save()
            messages.success(request, "Notification preferences saved.")
            return redirect("/admin-panel/settings?tab=notifications")

        if section == "account_profile":
            u = request.user
            u.first_name = P.get("first_name", u.first_name or "").strip() or None
            u.last_name  = P.get("last_name",  u.last_name  or "").strip() or None
            u.mobile     = P.get("mobile",     u.mobile     or "").strip() or None
            u.save(update_fields=["first_name", "last_name", "mobile"])
            messages.success(request, "Profile updated.")
            return redirect("/admin-panel/settings?tab=account")

        if section == "account_password":
            cur = P.get("current_password", "")
            new = P.get("new_password", "").strip()
            cnf = P.get("confirm_password", "").strip()
            if not request.user.check_password(cur):
                messages.error(request, "Current password is incorrect.")
            elif len(new) < 6:
                messages.error(request, "New password must be at least 6 characters.")
            elif new != cnf:
                messages.error(request, "Passwords do not match.")
            else:
                request.user.set_password(new)
                request.user.save()
                messages.success(request, "Password changed. Please log in again.")
                return redirect("/admin-login")
            return redirect("/admin-panel/settings?tab=account")

        if section == "payments":
            known = ["key_id", "key_secret", "webhook_secret", "environment", "enabled"]
            secret_keys = {"key_secret", "webhook_secret"}
            for key in known:
                value = P.get(key, "").strip()
                IntegrationConfig.set_value("RAZORPAY", key, value, key in secret_keys)
            messages.success(request, "Razorpay settings saved.")
            return redirect("/admin-panel/settings?tab=payments")

        if section == "integrations":
            known = ["cloud_name", "api_key", "api_secret", "folder",
                     "upload_preset", "max_size_mb", "allowed_types", "enabled"]
            secret_keys = {"api_secret"}
            for key in known:
                value = P.get(key, "").strip()
                IntegrationConfig.set_value("CLOUDINARY", key, value, key in secret_keys)
            messages.success(request, "Cloudinary settings saved.")
            return redirect("/admin-panel/settings?tab=integrations")

        if section == "team_add":
            email    = P.get("email", "").strip().lower()
            fname    = P.get("first_name", "").strip()
            lname    = P.get("last_name", "").strip()
            password = P.get("password", "").strip()
            if not email or not password:
                messages.error(request, "Email and password are required.")
            elif _User.objects.filter(email=email).exists():
                messages.error(request, "An account with this email already exists.")
            else:
                _User.objects.create_superuser(
                    email=email, password=password,
                    first_name=fname or None, last_name=lname or None,
                )
                messages.success(request, f"Team member {email} added.")
            return redirect("/admin-panel/settings?tab=team")

        if section == "team_toggle":
            uid = P.get("user_id")
            member = get_object_or_404(_User, pk=uid, user_type=_UT.ADMIN)
            if member.pk == request.user.pk:
                messages.error(request, "You cannot disable your own account.")
            else:
                from accounts.models import Status as _St
                member.status = _St.INACTIVE if member.status == _St.ACTIVE else _St.ACTIVE
                member.save(update_fields=["status"])
                messages.success(request, f"Account {'disabled' if member.status == _St.INACTIVE else 'enabled'}.")
            return redirect("/admin-panel/settings?tab=team")

        messages.warning(request, "Unknown settings section.")
        return redirect(f"/admin-panel/settings?tab={tab}")

    # ── GET ──
    cfg = SiteSettings.get()

    def _ic(integration, key, default=""):
        return IntegrationConfig.get(integration, key, default)

    rz = {
        "key_id":         _ic("RAZORPAY", "key_id"),
        "key_secret":     _ic("RAZORPAY", "key_secret"),
        "webhook_secret": _ic("RAZORPAY", "webhook_secret"),
        "environment":    _ic("RAZORPAY", "environment", "test"),
        "enabled":        _ic("RAZORPAY", "enabled", "false") == "true",
    }
    cl = {
        "cloud_name":    _ic("CLOUDINARY", "cloud_name"),
        "api_key":       _ic("CLOUDINARY", "api_key"),
        "api_secret":    _ic("CLOUDINARY", "api_secret"),
        "folder":        _ic("CLOUDINARY", "folder", "products"),
        "upload_preset": _ic("CLOUDINARY", "upload_preset"),
        "max_size_mb":   _ic("CLOUDINARY", "max_size_mb", "5"),
        "allowed_types": _ic("CLOUDINARY", "allowed_types", "jpg,jpeg,png,webp"),
        "enabled":       _ic("CLOUDINARY", "enabled", "false") == "true",
    }
    team = list(_User.objects.filter(user_type=_UT.ADMIN).order_by("first_name", "email"))

    return render(request, "panel/settings.html", _ctx(
        "/admin-panel/settings",
        cfg=cfg, tab=tab, rz=rz, cl=cl, team=team,
    ))


# --------------------------------------------------------------------------- #
# Archive management
# --------------------------------------------------------------------------- #
@admin_required
def archived_products(request):
    if request.method == "POST":
        action = request.POST.get("bulk_action", "")
        ids = request.POST.getlist("product_ids")
        if ids:
            if action == "restore":
                Product.objects.filter(pk__in=ids).update(
                    status=Status.ACTIVE, position=next_position(Product)
                )
                messages.success(request, f"Restored {len(ids)} product(s).")
            elif action == "delete":
                Product.objects.filter(pk__in=ids).delete()
                repack_positions(Product)
                messages.success(request, f"Permanently deleted {len(ids)} product(s).")
        return redirect("/admin-panel/products/archived")

    q = request.GET.get("q", "").strip()
    qs = (Product.objects.filter(status=Status.INACTIVE)
          .select_related("category").prefetch_related("variants")
          .order_by("-updated_at"))
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    return render(request, "panel/archived_products.html", _ctx(
        "/admin-panel/products/archived",
        products=qs, q=q, total=qs.count(),
    ))


@admin_required
def product_restore(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/products?show=archived")
    product = get_object_or_404(Product, pk=pk)
    product.status = Status.ACTIVE
    product.position = next_position(Product)
    product.save(update_fields=["status", "position"])
    messages.success(request, f"'{product.name}' restored to active.")
    return redirect("/admin-panel/products?show=archived")


# --------------------------------------------------------------------------- #
# Product import (CSV / Excel)
# --------------------------------------------------------------------------- #
@admin_required
def product_import(request):
    if request.method == "GET":
        field_ref = [
            ("name",     "Required",  "Product name — used to match existing products"),
            ("category", "Required",  "Must match an existing category name exactly"),
            ("sku",      "Optional",  "Product code; auto-generated if blank"),
            ("variant",  "Required",  "Variant label e.g. '100g', '500ml', 'Pack of 12'"),
            ("mrp",      "Optional",  "Original / strike-through price (number)"),
            ("price",    "Required",  "Selling price (number)"),
            ("stock",    "Optional",  "Quantity in stock, default 0"),
            ("badge",    "Optional",  "Label shown on product card e.g. 'Bestseller'"),
            ("status",   "Optional",  "ACTIVE or INACTIVE (default ACTIVE)"),
        ]
        return render(request, "panel/product_import.html",
                      _ctx("/admin-panel/products/import",
                           categories=Category.objects.order_by("position"),
                           field_ref=field_ref))

    uploaded = request.FILES.get("import_file")
    if not uploaded:
        messages.error(request, "No file uploaded.")
        return redirect("/admin-panel/products/import")

    ext = uploaded.name.rsplit(".", 1)[-1].lower()
    rows = []
    errors = []

    try:
        if ext == "csv":
            import io
            content = uploaded.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower().replace(" ", "_") if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in excel_row):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in excel_row])))
        else:
            messages.error(request, "Unsupported format. Upload a .csv or .xlsx file.")
            return redirect("/admin-panel/products/import")
    except Exception as e:
        messages.error(request, f"File parse error: {e}")
        return redirect("/admin-panel/products/import")

    # Normalise header names
    def _get(row, *keys):
        for k in keys:
            if k in row:
                return (row[k] or "").strip()
        return ""

    created_p = created_v = updated_v = 0
    cat_cache = {}

    for i, row in enumerate(rows, start=2):
        name         = _get(row, "name", "product_name", "product")
        cat_name     = _get(row, "category", "category_name")
        sku          = _get(row, "sku", "code", "product_sku")
        variant_name = _get(row, "variant", "variant_name", "size")
        mrp          = _get(row, "mrp", "mrp_price", "original_price")
        price        = _get(row, "price", "selling_price", "sale_price")
        stock        = _get(row, "stock", "quantity", "qty")
        badge        = _get(row, "badge", "label")
        status_val   = _get(row, "status").upper() or Status.ACTIVE

        if not name or not cat_name or not variant_name or not price:
            errors.append(f"Row {i}: missing required fields (name, category, variant, price).")
            continue

        # Find/create category
        if cat_name not in cat_cache:
            try:
                cat_cache[cat_name] = Category.objects.get(name__iexact=cat_name)
            except Category.DoesNotExist:
                errors.append(f"Row {i}: category '{cat_name}' not found — skipped.")
                continue
        category = cat_cache[cat_name]

        # Find/create product
        try:
            mrp_f   = float(mrp or 0)
            price_f = float(price)
            stock_i = int(float(stock or 0))
        except ValueError:
            errors.append(f"Row {i}: invalid numeric value.")
            continue

        if sku:
            product, new_p = Product.objects.get_or_create(
                code=sku,
                defaults={"name": name, "slug": slugify(name),
                          "category": category, "status": status_val or Status.ACTIVE,
                          "image": "/seed/placeholder.jpg", "badge": badge or None},
            )
        else:
            product, new_p = Product.objects.get_or_create(
                name=name, category=category,
                defaults={"code": gen_code("PRD"), "slug": slugify(name),
                          "status": status_val or Status.ACTIVE,
                          "image": "/seed/placeholder.jpg", "badge": badge or None},
            )
        if new_p:
            created_p += 1

        # Find/create variant
        va_code_val = _get(row, "variant_sku", "va_code") or gen_code("SKU")
        v_qs = ProductVariant.objects.filter(product=product, variant__iexact=variant_name)
        if v_qs.exists():
            v = v_qs.first()
            v.mrp_price = mrp_f
            v.selling_price = price_f
            v.stock = stock_i
            from store.models import StockThreshold
            threshold = StockThreshold.get_for_product(product)
            v.stock_status = threshold.compute_status(stock_i)
            v.save(update_fields=["mrp_price", "selling_price", "stock", "stock_status"])
            updated_v += 1
        else:
            from store.models import StockThreshold
            threshold = StockThreshold.get_for_product(product)
            ProductVariant.objects.create(
                product=product, variant=variant_name,
                va_code=va_code_val, mrp_price=mrp_f, selling_price=price_f,
                stock=stock_i, stock_status=threshold.compute_status(stock_i),
                status=Status.ACTIVE,
            )
            created_v += 1

    msg = f"Import complete: {created_p} product(s) created, {created_v} variant(s) added, {updated_v} variant(s) updated."
    if errors:
        msg += f" {len(errors)} row(s) skipped."
        messages.warning(request, msg)
        for e in errors[:10]:
            messages.warning(request, e)
    else:
        messages.success(request, msg)
    return redirect("/admin-panel/products")


# --------------------------------------------------------------------------- #
# Media / image upload (Cloudinary only — no local fallback)
# --------------------------------------------------------------------------- #
@admin_required
def media_upload(request):
    import json as _json

    if request.method != "POST":
        return HttpResponse(status=405)

    f = request.FILES.get("file")
    if not f:
        return HttpResponse(_json.dumps({"error": "No file provided"}),
                            content_type="application/json", status=400)

    # Load limits from Cloudinary config (fallback to safe defaults)
    from store.models import IntegrationConfig as _IC
    try:
        _max_mb = float(_IC.get("CLOUDINARY", "max_size_mb", "5") or "5")
    except ValueError:
        _max_mb = 5.0
    _raw_types = _IC.get("CLOUDINARY", "allowed_types", "jpg,jpeg,png,webp,gif")
    _ext_list = [t.strip().lower().lstrip(".") for t in _raw_types.split(",") if t.strip()]
    _mime_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                 "webp": "image/webp", "gif": "image/gif", "svg": "image/svg+xml"}
    allowed_mimes = {_mime_map.get(e, f"image/{e}") for e in _ext_list} or {
        "image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"}

    if f.content_type not in allowed_mimes:
        return HttpResponse(
            _json.dumps({"error": f"Only {_raw_types} images are allowed"}),
            content_type="application/json", status=400)

    if f.size > _max_mb * 1024 * 1024:
        return HttpResponse(_json.dumps({"error": f"File too large — maximum {_max_mb:.0f} MB"}),
                            content_type="application/json", status=400)

    import uuid as _uuid
    ext = f.name.rsplit(".", 1)[-1].lower() if "." in f.name else "jpg"
    unique_name = f"{_uuid.uuid4().hex}.{ext}"
    file_bytes = f.read()

    from core.cloudinary_storage import upload_file as cl_upload
    url, err = cl_upload(file_bytes, unique_name, f.content_type)
    if url:
        return HttpResponse(_json.dumps({"url": url}),
                            content_type="application/json")

    return HttpResponse(
        _json.dumps({"error": err or "Cloud storage is not configured. Please enable Cloudinary in Settings → Integrations."}),
        content_type="application/json", status=503,
    )


# --------------------------------------------------------------------------- #
# Integrations module
# --------------------------------------------------------------------------- #
@admin_required
def integrations_view(request):
    from store.models import IntegrationConfig

    if request.method == "POST":
        integration = request.POST.get("integration", "").upper()
        if integration not in ("RAZORPAY", "CLOUDINARY", "EMAIL", "SMS"):
            messages.error(request, "Unknown integration.")
            return redirect("/admin-panel/integrations")

        # Save all posted keys for this integration
        known_keys = {
            "RAZORPAY":   ["key_id", "key_secret", "webhook_secret", "environment", "enabled"],
            "CLOUDINARY": ["cloud_name", "api_key", "api_secret", "folder", "upload_preset",
                           "max_size_mb", "allowed_types", "enabled"],
            "EMAIL":      ["host", "port", "username", "password", "use_tls", "from_email", "enabled"],
            "SMS":        ["provider", "api_key", "account_sid", "sender_id", "enabled"],
        }
        secret_keys = {"key_secret", "webhook_secret", "api_secret", "password", "api_key"}

        for key in known_keys[integration]:
            value = request.POST.get(key, "").strip()
            IntegrationConfig.set_value(
                integration=integration, key=key, value=value,
                is_secret=(key in secret_keys),
            )
        messages.success(request, f"{integration.title()} settings saved.")
        return redirect("/admin-panel/integrations")

    def cfg(integration, key, default=""):
        return IntegrationConfig.get(integration, key, default)

    ctx = {
        "rz": {
            "key_id":         cfg("RAZORPAY", "key_id"),
            "key_secret":     cfg("RAZORPAY", "key_secret"),
            "webhook_secret": cfg("RAZORPAY", "webhook_secret"),
            "environment":    cfg("RAZORPAY", "environment", "test"),
            "enabled":        cfg("RAZORPAY", "enabled", "false") == "true",
        },
        "cl": {
            "cloud_name":     cfg("CLOUDINARY", "cloud_name"),
            "api_key":        cfg("CLOUDINARY", "api_key"),
            "api_secret":     cfg("CLOUDINARY", "api_secret"),
            "folder":         cfg("CLOUDINARY", "folder", "products"),
            "upload_preset":  cfg("CLOUDINARY", "upload_preset", ""),
            "max_size_mb":    cfg("CLOUDINARY", "max_size_mb", "5"),
            "allowed_types":  cfg("CLOUDINARY", "allowed_types", "jpg,jpeg,png,webp,gif"),
            "enabled":        cfg("CLOUDINARY", "enabled", "false") == "true",
        },
        "em": {
            "host":       cfg("EMAIL", "host", ""),
            "port":       cfg("EMAIL", "port", "587"),
            "username":   cfg("EMAIL", "username", ""),
            "password":   cfg("EMAIL", "password", ""),
            "use_tls":    cfg("EMAIL", "use_tls", "true") == "true",
            "from_email": cfg("EMAIL", "from_email", ""),
            "enabled":    cfg("EMAIL", "enabled", "false") == "true",
        },
        "sm": {
            "provider":    cfg("SMS", "provider", "fast2sms"),
            "api_key":     cfg("SMS", "api_key", ""),
            "account_sid": cfg("SMS", "account_sid", ""),
            "sender_id":   cfg("SMS", "sender_id", ""),
            "enabled":     cfg("SMS", "enabled", "false") == "true",
        },
    }
    return render(request, "panel/integrations.html",
                  _ctx("/admin-panel/integrations", **ctx))


@admin_required
def integration_test(request, integration):
    import json as _json
    if request.method != "POST":
        return HttpResponse(status=405)
    integration = integration.upper()

    if integration == "CLOUDINARY":
        from core.cloudinary_storage import test_connection
        ok, msg = test_connection()
    elif integration == "RAZORPAY":
        ok, msg = _test_razorpay()
    elif integration == "EMAIL":
        ok, msg = _test_email()
    elif integration == "SMS":
        ok, msg = _test_sms()
    else:
        ok, msg = False, "Unknown integration."

    return HttpResponse(_json.dumps({"ok": ok, "message": msg}),
                        content_type="application/json")


def _test_razorpay():
    from store.models import IntegrationConfig
    key_id     = IntegrationConfig.get("RAZORPAY", "key_id")
    key_secret = IntegrationConfig.get("RAZORPAY", "key_secret")
    if not key_id or not key_secret:
        return False, "Razorpay Key ID and Key Secret are required."
    try:
        import razorpay
        client = razorpay.Client(auth=(key_id, key_secret))
        # A lightweight API call to verify credentials
        client.order.all({"count": 1})
        return True, f"Razorpay connection successful (environment: {IntegrationConfig.get('RAZORPAY', 'environment', 'test')})"
    except ImportError:
        return False, "razorpay package not installed."
    except Exception as e:
        err = str(e)
        if "401" in err or "authentication" in err.lower():
            return False, "Authentication failed — check your Key ID and Key Secret."
        return False, f"Connection error: {err}"


def _test_email():
    from store.models import IntegrationConfig
    host       = IntegrationConfig.get("EMAIL", "host", "")
    port       = IntegrationConfig.get("EMAIL", "port", "587")
    username   = IntegrationConfig.get("EMAIL", "username", "")
    password   = IntegrationConfig.get("EMAIL", "password", "")
    use_tls    = IntegrationConfig.get("EMAIL", "use_tls", "true") == "true"
    from_email = IntegrationConfig.get("EMAIL", "from_email", "") or username
    if not host or not username:
        return False, "SMTP host and username are required."
    try:
        from django.core.mail import get_connection
        conn = get_connection(
            backend="django.core.mail.backends.smtp.EmailBackend",
            host=host, port=int(port or "587"),
            username=username, password=password,
            use_tls=use_tls, fail_silently=False,
        )
        conn.open()
        conn.close()
        return True, f"SMTP connection successful ({username} via {host}:{port})"
    except Exception as e:
        err = str(e)
        if "authentication" in err.lower() or "535" in err or "534" in err:
            return False, "Authentication failed — check your username and password (or App Password for Gmail)."
        if "timed out" in err.lower() or "connection refused" in err.lower():
            return False, f"Cannot reach {host}:{port} — check host, port, and firewall."
        return False, f"SMTP error: {err}"


def _test_sms():
    from store.models import IntegrationConfig
    provider    = IntegrationConfig.get("SMS", "provider", "")
    api_key     = IntegrationConfig.get("SMS", "api_key", "")
    account_sid = IntegrationConfig.get("SMS", "account_sid", "")
    if not provider or not api_key:
        return False, "Provider and API Key are required."
    try:
        import requests as _req

        if provider == "fast2sms":
            resp = _req.get(
                "https://www.fast2sms.com/dev/wallet",
                headers={"authorization": api_key, "cache-control": "no-cache"},
                timeout=10,
            )
            data = resp.json()
            if data.get("return"):
                bal = data.get("wallet", {}).get("wallet", "?")
                return True, f"Fast2SMS connected. Wallet balance: ₹{bal}"
            msg = data.get("message", "Authentication failed")
            return False, (msg[0] if isinstance(msg, list) else str(msg))

        elif provider == "msg91":
            resp = _req.get(
                "https://api.msg91.com/api/balance.php",
                params={"authkey": api_key, "type": "json"},
                timeout=10,
            )
            data = resp.json()
            if "Balance" in data:
                return True, f"MSG91 connected. SMS balance: {data['Balance']}"
            return False, data.get("message", "Authentication failed")

        elif provider == "twilio":
            if not account_sid:
                return False, "Account SID is required for Twilio."
            import base64
            creds = base64.b64encode(f"{account_sid}:{api_key}".encode()).decode()
            resp = _req.get(
                f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}.json",
                headers={"Authorization": f"Basic {creds}"},
                timeout=10,
            )
            if resp.status_code == 200:
                data = resp.json()
                return True, f"Twilio connected. Account: {data.get('friendly_name', account_sid)}"
            return False, resp.json().get("message", "Authentication failed")

        return False, f"Unknown provider: {provider}"
    except Exception as e:
        return False, f"Connection error: {e}"


# --------------------------------------------------------------------------- #
# Combo Packs
# --------------------------------------------------------------------------- #
@admin_required
def combo_list(request):
    from django.http import JsonResponse

    if request.method == "POST":
        action = request.POST.get("action", "")
        ids = request.POST.getlist("ids")
        if ids:
            if action == "delete":
                ComboPackage.objects.filter(pk__in=ids).delete()
                messages.success(request, f"Deleted {len(ids)} combo(s).")
            elif action == "publish":
                ComboPackage.objects.filter(pk__in=ids).update(status=ComboPackage.Status.ACTIVE)
                messages.success(request, f"Published {len(ids)} combo(s).")
            elif action == "draft":
                ComboPackage.objects.filter(pk__in=ids).update(status=ComboPackage.Status.DRAFT)
                messages.success(request, f"Moved {len(ids)} combo(s) to draft.")
        return redirect("/admin-panel/combos")

    qs = ComboPackage.objects.prefetch_related("items__variant", "images")

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))

    status_filter = request.GET.get("status", "")
    if status_filter:
        qs = qs.filter(status=status_filter)

    sort = request.GET.get("sort", "")
    qs = qs.order_by({
        "name":       "name",
        "price_asc":  "selling_price",
        "price_desc": "-selling_price",
        "orders":     "-orders_count",
        "newest":     "-created_at",
        "oldest":     "created_at",
    }.get(sort, "position"))

    all_qs = ComboPackage.objects.all()
    total_combos    = all_qs.count()
    active_combos   = all_qs.filter(status=ComboPackage.Status.ACTIVE).count()
    featured_combos = all_qs.filter(is_featured=True).count()
    total_orders    = all_qs.aggregate(s=Sum("orders_count"))["s"] or 0

    active_prefetched = ComboPackage.objects.filter(
        status=ComboPackage.Status.ACTIVE
    ).prefetch_related("items__variant")
    savings_pcts = [c.savings_pct for c in active_prefetched if c.savings_pct > 0]
    avg_savings = round(sum(savings_pcts) / len(savings_pcts), 1) if savings_pcts else 0

    combo_revenue = sum(
        c.selling_price * c.orders_count
        for c in all_qs
    )

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "panel/combos.html", {
        "combos":         page_obj,
        "page_obj":       page_obj,
        "filtered_total": qs.count(),
        "q":              q,
        "status_filter":  status_filter,
        "sort":           sort,
        "total_combos":   total_combos,
        "active_combos":  active_combos,
        "featured_combos": featured_combos,
        "total_orders":   total_orders,
        "avg_savings":    avg_savings,
        "combo_revenue":  combo_revenue,
        "status_choices": ComboPackage.Status.choices,
    })


@admin_required
def combo_variant_search(request):
    from django.http import JsonResponse
    q = request.GET.get("q", "").strip()
    if len(q) >= 2:
        variants = (ProductVariant.objects
                    .filter(
                        Q(product__name__icontains=q) | Q(va_code__icontains=q),
                        product__status=Status.ACTIVE, status=Status.ACTIVE,
                    )
                    .select_related("product")[:10])
    else:
        # Return popular variants as suggestions when no query
        variants = (ProductVariant.objects
                    .filter(product__status=Status.ACTIVE, status=Status.ACTIVE)
                    .select_related("product")
                    .order_by("product__position", "position")[:8])
    results = [{
        "id":           v.pk,
        "name":         v.product.name,
        "variant":      v.variant,
        "sku":          v.va_code,
        "price":        v.selling_price,
        "stock":        v.stock,
        "stock_status": v.stock_status,
        "image":        v.product.image or "",
    } for v in variants]
    return JsonResponse({"results": results})


@admin_required
def combo_edit(request, pk=None):
    combo = get_object_or_404(ComboPackage, pk=pk) if pk else None
    items  = list(combo.items.select_related("variant__product").order_by("position")) if combo else []
    images = list(combo.images.all()) if combo else []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if not name:
            messages.error(request, "Combo name is required.")
            return redirect(request.path)

        is_new = not combo
        if is_new:
            combo = ComboPackage()
            combo.code = gen_code("CMB")
        else:
            old_pos = combo.position

        combo.name              = name
        combo.slug              = slugify(name)
        combo.short_description = request.POST.get("short_description", "")
        combo.description       = request.POST.get("description", "")
        combo.badge_label       = request.POST.get("badge_label", "")
        combo.badge_style       = request.POST.get("badge_style", "")
        combo.tags              = request.POST.get("tags", "")
        combo.selling_price     = float(request.POST.get("selling_price", "0") or "0")
        mrp_raw                 = request.POST.get("mrp_price", "")
        combo.mrp_price         = float(mrp_raw) if mrp_raw else None
        combo.gst_rate          = float(request.POST.get("gst_rate", "5") or "5")
        combo.status            = request.POST.get("status", ComboPackage.Status.DRAFT)
        combo.is_featured       = request.POST.get("is_featured") == "1"
        combo.is_limited_time   = request.POST.get("is_limited_time") == "1"
        combo.is_cod_available  = request.POST.get("is_cod_available") == "1"
        combo.available_from    = request.POST.get("available_from") or None
        combo.available_until   = request.POST.get("available_until") or None
        combo.max_qty_per_order = int(request.POST.get("max_qty_per_order", "10") or "10")

        raw_pos = request.POST.get("position", "").strip()
        new_pos = int(raw_pos) if raw_pos.isdigit() else (
            next_position(ComboPackage) if is_new else old_pos
        )
        combo.position = new_pos

        base_slug = combo.slug
        counter = 1
        while ComboPackage.objects.filter(slug=combo.slug).exclude(pk=combo.pk or 0).exists():
            combo.slug = f"{base_slug}-{counter}"
            counter += 1

        if is_new:
            with transaction.atomic():
                if ComboPackage.objects.filter(position=new_pos).exists():
                    insert_at_position(ComboPackage, new_pos)
                combo.save()
        else:
            combo.save()
            if new_pos != old_pos:
                move_to_position(ComboPackage, combo.pk, new_pos, old_pos)

        variant_ids = request.POST.getlist("item_variant_id")
        quantities  = request.POST.getlist("item_qty")
        combo.items.all().delete()
        for i, (vid, qty) in enumerate(zip(variant_ids, quantities)):
            try:
                v = ProductVariant.objects.get(pk=int(vid))
                ComboItem.objects.create(combo=combo, variant=v,
                                         quantity=int(qty or 1), position=i)
            except (ProductVariant.DoesNotExist, ValueError):
                pass

        image_urls = request.POST.getlist("image_url")
        image_main = request.POST.get("image_main_idx", "0")
        combo.images.all().delete()
        for i, url in enumerate(image_urls):
            if url.strip():
                ComboImage.objects.create(
                    combo=combo, url=url.strip(),
                    is_main=(str(i) == image_main), position=i,
                )

        label = "published" if combo.status == ComboPackage.Status.ACTIVE else "saved as draft"
        messages.success(request, f'Combo "{combo.name}" {label}.')
        return redirect(f"/admin-panel/combos/{combo.pk}/edit")

    return render(request, "panel/combo_edit.html", {
        "combo":          combo,
        "combo_items":    items,
        "combo_images":   images,
        "badge_styles":   ComboPackage.BadgeStyle.choices,
        "gst_rates": [
            ("0",  "0% (Exempt)"),
            ("5",  "5% (Food Products)"),
            ("12", "12%"),
            ("18", "18%"),
            ("28", "28%"),
        ],
        "status_choices": ComboPackage.Status.choices,
        "next_pos":       next_position(ComboPackage) if not combo else None,
    })


@admin_required
def combo_delete(request, pk):
    combo = get_object_or_404(ComboPackage, pk=pk)
    name = combo.name
    combo.delete()
    repack_positions(ComboPackage)
    messages.success(request, f'Combo "{name}" deleted.')
    return redirect("/admin-panel/combos")


# --------------------------------------------------------------------------- #
# Reviews module
# --------------------------------------------------------------------------- #

AVATAR_STYLES = [
    "bg-brand-100 text-brand-700",
    "bg-blue-100 text-blue-700",
    "bg-purple-100 text-purple-700",
    "bg-orange-100 text-orange-700",
    "bg-teal-100 text-teal-700",
    "bg-rose-100 text-rose-700",
    "bg-indigo-100 text-indigo-700",
    "bg-amber-100 text-amber-700",
]


@admin_required
def reviews_list(request):
    from datetime import timedelta
    from django.http import JsonResponse

    # Bulk POST actions
    if request.method == "POST":
        action = request.POST.get("action", "")
        ids = request.POST.getlist("ids")
        if ids:
            affected = list(Review.objects.filter(pk__in=ids).select_related("user", "product"))
            if action == "delete":
                Review.objects.filter(pk__in=ids).delete()
                messages.success(request, f"Deleted {len(ids)} review(s).")
            elif action == "approve":
                Review.objects.filter(pk__in=ids).update(status=ReviewStatus.APPROVED, is_flagged=False)
                for r in affected:
                    Notification.objects.create(
                        user=r.user,
                        title="Your review has been approved",
                        message=f"Your review for \"{r.product.name}\" is now live. Thank you!",
                    )
                messages.success(request, f"Approved {len(ids)} review(s).")
            elif action == "reject":
                Review.objects.filter(pk__in=ids).update(status=ReviewStatus.REJECTED)
                for r in affected:
                    Notification.objects.create(
                        user=r.user,
                        title="Your review was not approved",
                        message=f"Unfortunately your review for \"{r.product.name}\" did not meet our guidelines and was not published.",
                    )
                messages.success(request, f"Rejected {len(ids)} review(s).")
            elif action == "unflag":
                Review.objects.filter(pk__in=ids).update(is_flagged=False, flag_reason="")
                messages.success(request, f"Unflagged {len(ids)} review(s).")
        return redirect("/admin-panel/reviews")

    # Period
    try:
        period = int(request.GET.get("period", "30") or "30")
    except ValueError:
        period = 30

    now = timezone.now()
    if period > 0:
        since = now - timedelta(days=period)
        period_qs = Review.objects.filter(created_at__gte=since)
    else:
        period_qs = Review.objects.all()
        since = None

    approved_qs = period_qs.filter(status=ReviewStatus.APPROVED)
    total_count = approved_qs.count()
    avg_rating_val = approved_qs.aggregate(a=Avg("rating"))["a"] or 0
    avg_rating = round(avg_rating_val, 1)

    # Compare to previous period
    if period > 0 and since is not None:
        prev_since = since - timedelta(days=period)
        prev_avg = (
            Review.objects.filter(
                created_at__gte=prev_since, created_at__lt=since,
                status=ReviewStatus.APPROVED
            ).aggregate(a=Avg("rating"))["a"] or 0
        )
        avg_diff = round(avg_rating - prev_avg, 1)
    else:
        avg_diff = 0

    # Rating breakdown (5→1 descending for display)
    breakdown = {}
    for i in range(1, 6):
        breakdown[i] = approved_qs.filter(rating=i).count()
    breakdown_list = [
        {
            "star": i,
            "count": breakdown[i],
            "pct": round(breakdown[i] / total_count * 100) if total_count else 0,
        }
        for i in range(5, 0, -1)
    ]

    # Health metrics
    all_period = period_qs.count()
    replied_count = period_qs.exclude(reply="").count()
    response_rate = round(replied_count / all_period * 100) if all_period else 0
    pending_replies = period_qs.filter(status=ReviewStatus.APPROVED, reply="").count()
    flagged_count = period_qs.filter(is_flagged=True).count()

    # Avg response time (hours)
    replied_timing = list(
        period_qs.filter(replied_at__isnull=False).values("created_at", "replied_at")
    )
    if replied_timing:
        total_secs = sum(
            (r["replied_at"] - r["created_at"]).total_seconds() for r in replied_timing
        )
        avg_response_hrs = round(total_secs / len(replied_timing) / 3600, 1)
    else:
        avg_response_hrs = 0

    # Pending count (always global, not period-restricted)
    pending_count = Review.objects.filter(status=ReviewStatus.PENDING).count()

    # Filters
    q = request.GET.get("q", "").strip()
    product_filter = request.GET.get("product", "")
    rating_filter = request.GET.get("rating", "")
    reply_filter = request.GET.get("replied", "")
    status_filter = request.GET.get("status", "")
    sort = request.GET.get("sort", "-created_at")

    qs = period_qs.select_related("user", "product").prefetch_related("images")

    if q:
        qs = qs.filter(
            Q(title__icontains=q) | Q(comment__icontains=q) |
            Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) |
            Q(product__name__icontains=q)
        )
    if product_filter:
        qs = qs.filter(product_id=product_filter)
    if rating_filter:
        try:
            qs = qs.filter(rating=int(rating_filter))
        except ValueError:
            pass
    if reply_filter == "replied":
        qs = qs.exclude(reply="")
    elif reply_filter == "unreplied":
        qs = qs.filter(reply="")
    elif reply_filter == "flagged":
        qs = qs.filter(is_flagged=True)

    if status_filter in ("PENDING", "APPROVED", "REJECTED", "FLAGGED"):
        if status_filter == "FLAGGED":
            qs = qs.filter(is_flagged=True)
        else:
            qs = qs.filter(status=status_filter)

    sort_map = {
        "-created_at": "-created_at",
        "created_at": "created_at",
        "-rating": "-rating",
        "rating": "rating",
        "-helpful_count": "-helpful_count",
    }
    qs = qs.order_by(sort_map.get(sort, "-created_at"))

    products_with_reviews = (
        Product.objects.filter(reviews__isnull=False).distinct().order_by("name")
    )

    paginator = Paginator(qs, 7)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    for r in page_obj:
        r.avatar_style = AVATAR_STYLES[r.user_id % len(AVATAR_STYLES)]
        r.initials = (
            ((r.user.first_name or "")[:1] + (r.user.last_name or "")[:1]).upper()
            or (r.user.email or "?")[:1].upper()
        )

    return render(request, "panel/reviews.html", {
        "active_path": "/admin-panel/reviews",
        "reviews": page_obj,
        "page_obj": page_obj,
        "filtered_total": qs.count(),
        "period": period,
        "avg_rating": avg_rating,
        "avg_diff": avg_diff,
        "total_count": total_count,
        "breakdown": breakdown,
        "breakdown_list": breakdown_list,
        "response_rate": response_rate,
        "pending_replies": pending_replies,
        "pending_count": pending_count,
        "flagged_count": flagged_count,
        "avg_response_hrs": avg_response_hrs,
        "q": q,
        "product_filter": product_filter,
        "rating_filter": rating_filter,
        "reply_filter": reply_filter,
        "status_filter": status_filter,
        "sort": sort,
        "products_with_reviews": products_with_reviews,
    })


@admin_required
def review_reply(request, pk):
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    review = get_object_or_404(Review, pk=pk)
    reply_text = request.POST.get("reply", "").strip()
    if not reply_text:
        return JsonResponse({"error": "Reply text cannot be empty."}, status=400)

    review.reply = reply_text
    review.replied_at = timezone.now()
    review.save(update_fields=["reply", "replied_at"])

    return JsonResponse({"ok": True, "reply": reply_text})


@admin_required
def review_approve(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/reviews")
    review = get_object_or_404(Review, pk=pk)
    review.status = ReviewStatus.APPROVED
    review.is_flagged = False
    review.save(update_fields=["status", "is_flagged"])
    Notification.objects.create(
        user=review.user,
        title="Your review has been approved",
        message=f"Your review for \"{review.product.name}\" is now live. Thank you!",
    )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        from django.http import JsonResponse
        return JsonResponse({"ok": True})
    messages.success(request, "Review approved.")
    return redirect(request.POST.get("next", "/admin-panel/reviews"))


@admin_required
def review_reject(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/reviews")
    review = get_object_or_404(Review, pk=pk)
    review.status = ReviewStatus.REJECTED
    review.save(update_fields=["status"])
    Notification.objects.create(
        user=review.user,
        title="Your review was not approved",
        message=f"Your review for \"{review.product.name}\" did not meet our guidelines and was not published.",
    )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        from django.http import JsonResponse
        return JsonResponse({"ok": True})
    messages.success(request, "Review rejected.")
    return redirect(request.POST.get("next", "/admin-panel/reviews"))


@admin_required
def review_flag(request, pk):
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    review = get_object_or_404(Review, pk=pk)
    reason = request.POST.get("reason", "").strip()
    if review.is_flagged:
        review.is_flagged = False
        review.flag_reason = ""
    else:
        review.is_flagged = True
        review.flag_reason = reason
        Notification.objects.create(
            user=review.user,
            title="Your review has been flagged",
            message=f"Your review for \"{review.product.name}\" has been flagged for the following reason: {reason or 'Policy violation'}. It is no longer publicly visible.",
        )
    review.save(update_fields=["is_flagged", "flag_reason"])
    return JsonResponse({"ok": True, "flagged": review.is_flagged})


@admin_required
def review_delete(request, pk):
    if request.method != "POST":
        return redirect("/admin-panel/reviews")
    review = get_object_or_404(Review, pk=pk)
    review.delete()
    messages.success(request, "Review deleted.")
    return redirect(request.POST.get("next", "/admin-panel/reviews"))


@admin_required
def review_export(request):
    period = int(request.GET.get("period", "30") or "0")
    now = timezone.now()
    qs = Review.objects.select_related("user", "product").order_by("-created_at")
    if period > 0:
        from datetime import timedelta
        qs = qs.filter(created_at__gte=now - timedelta(days=period))

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reviews.csv"'
    writer = csv.writer(response)
    writer.writerow([
        "ID", "Customer", "Email", "Product", "Rating", "Title",
        "Comment", "Reply", "Status", "Verified", "Flagged",
        "Helpful Count", "Date",
    ])
    for r in qs:
        writer.writerow([
            r.pk,
            f"{r.user.first_name} {r.user.last_name}".strip() or r.user.email,
            r.user.email,
            r.product.name,
            r.rating,
            r.title,
            r.comment,
            r.reply,
            r.status,
            r.is_verified,
            r.is_flagged,
            r.helpful_count,
            r.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return response


# --------------------------------------------------------------------------- #
# Banners
# --------------------------------------------------------------------------- #

@admin_required
def banners_list(request):
    if request.method == "POST":
        action = request.POST.get("action", "")
        ids = request.POST.getlist("ids")
        if ids:
            if action == "delete":
                Banner.objects.filter(pk__in=ids).delete()
                messages.success(request, f"Deleted {len(ids)} banner(s).")
            elif action == "activate":
                Banner.objects.filter(pk__in=ids).update(status=Status.ACTIVE)
            elif action == "deactivate":
                Banner.objects.filter(pk__in=ids).update(status=Status.INACTIVE)
        return redirect("/admin-panel/banners")

    q = request.GET.get("q", "").strip()
    type_filter = request.GET.get("type", "")
    status_filter = request.GET.get("status", "")

    qs = Banner.objects.all()
    if q:
        qs = qs.filter(name__icontains=q)
    if type_filter:
        qs = qs.filter(type=type_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)

    total   = Banner.objects.count()
    active  = Banner.objects.filter(status=Status.ACTIVE).count()
    by_type = {t: Banner.objects.filter(type=t).count() for t, _ in BannerType.choices}

    return render(request, "panel/banners.html", {
        "active_path": "/admin-panel/banners",
        "banners": qs,
        "total": total,
        "active_count": active,
        "inactive_count": total - active,
        "by_type": by_type,
        "type_choices": BannerType.choices,
        "q": q,
        "type_filter": type_filter,
        "status_filter": status_filter,
    })


@admin_required
def banner_edit(request, pk=None):
    banner = get_object_or_404(Banner, pk=pk) if pk else None

    if request.method == "POST":
        name        = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        image       = request.POST.get("image", "").strip()
        video_url   = request.POST.get("video_url", "").strip()
        btype       = request.POST.get("type", BannerType.HOME_BANNER)
        try:
            position = int(request.POST.get("position", 1) or 1)
        except ValueError:
            position = 1
        status_val  = request.POST.get("status", Status.ACTIVE)

        if not name:
            messages.error(request, "Banner name is required.")
            return redirect(request.path)

        if banner:
            old_pos = banner.position
            banner.name        = name
            banner.description = description
            banner.image       = image
            banner.video_url   = video_url or None
            banner.type        = btype
            banner.position    = position
            banner.status      = status_val
            banner.save()
            if position != old_pos:
                move_to_position(Banner, banner.pk, position, old_pos)
            messages.success(request, f'Banner "{name}" updated.')
        else:
            with transaction.atomic():
                if Banner.objects.filter(position=position).exists():
                    insert_at_position(Banner, position)
                Banner.objects.create(
                    name=name, description=description, image=image,
                    video_url=video_url or None, type=btype,
                    position=position, status=status_val,
                )
            messages.success(request, f'Banner "{name}" created.')
        return redirect("/admin-panel/banners")

    return render(request, "panel/banner_form.html", {
        "active_path": "/admin-panel/banners",
        "banner": banner,
        "type_choices": BannerType.choices,
        "status_choices": Status.choices,
        "next_pos": next_position(Banner) if not banner else None,
    })


@admin_required
def banner_delete(request, pk):
    banner = get_object_or_404(Banner, pk=pk)
    name = banner.name
    banner.delete()
    repack_positions(Banner)
    messages.success(request, f'Banner "{name}" deleted.')
    return redirect("/admin-panel/banners")


@admin_required
def banner_toggle(request, pk):
    from django.http import JsonResponse
    banner = get_object_or_404(Banner, pk=pk)
    banner.status = Status.INACTIVE if banner.status == Status.ACTIVE else Status.ACTIVE
    banner.save(update_fields=["status"])
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"ok": True, "status": banner.status})
    return redirect("/admin-panel/banners")


# --------------------------------------------------------------------------- #
# Coupons / Offers
# --------------------------------------------------------------------------- #

def _coupon_classify(coupons_qs, status_filter, channel_filter, q):
    """Evaluate all coupons, attach revenue + status_tag, apply filters."""
    from django.db.models import Sum as _Sum
    qs = coupons_qs.annotate(revenue=_Sum("coupon_orders__grand_total"))
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(description__icontains=q))
    all_c = list(qs)
    if channel_filter:
        all_c = [c for c in all_c if channel_filter.upper() in c.channel_list]
    if status_filter in ("active", "scheduled", "expired", "inactive"):
        tgt = status_filter.capitalize()
        all_c = [c for c in all_c if c.status_tag == tgt]
    return all_c


@admin_required
def coupons_list(request):
    if request.method == "POST":
        action = request.POST.get("action", "")
        ids    = request.POST.getlist("ids")
        if ids:
            if action == "delete":
                Coupon.objects.filter(pk__in=ids).delete()
                messages.success(request, f"Deleted {len(ids)} coupon(s).")
            elif action == "activate":
                Coupon.objects.filter(pk__in=ids).update(is_active=True)
            elif action == "deactivate":
                Coupon.objects.filter(pk__in=ids).update(is_active=False)
        return redirect("/admin-panel/coupons")

    q              = request.GET.get("q", "").strip()
    status_filter  = request.GET.get("status", "").lower()
    channel_filter = request.GET.get("channel", "")
    sort           = request.GET.get("sort", "-created_at")
    sort_map = {"-created_at":"-created_at","created_at":"created_at",
                "-used_count":"-used_count","code":"code"}
    base_qs = Coupon.objects.all().order_by(sort_map.get(sort, "-created_at"))

    all_coupons = _coupon_classify(base_qs, status_filter, channel_filter, q)

    # Stats
    all_for_stats = list(Coupon.objects.all())
    stat_active    = sum(1 for c in all_for_stats if c.status_tag == "Active")
    stat_scheduled = sum(1 for c in all_for_stats if c.status_tag == "Scheduled")
    stat_expired   = sum(1 for c in all_for_stats if c.status_tag == "Expired")

    now = timezone.now()
    month_ago = now - __import__("datetime").timedelta(days=30)
    revenue_30d = Order.objects.filter(
        coupon__isnull=False, created_at__gte=month_ago
    ).aggregate(s=Sum("grand_total"))["s"] or 0

    paginator = Paginator(all_coupons, 10)
    page_obj  = paginator.get_page(request.GET.get("page", 1))

    def _chip(val, label):
        active_cls = "bg-brand-600 text-white"
        idle_cls   = "text-gray-600 hover:bg-gray-100"
        return (val, label, active_cls if status_filter == val else idle_cls)

    status_chips = [
        _chip("",           "All"),
        _chip("active",     "Active"),
        _chip("scheduled",  "Scheduled"),
        _chip("expired",    "Expired"),
    ]

    return render(request, "panel/coupons.html", {
        "active_path":    "/admin-panel/coupons",
        "coupons":        page_obj,
        "page_obj":       page_obj,
        "filtered_total": len(all_coupons),
        "stat_active":    stat_active,
        "stat_scheduled": stat_scheduled,
        "stat_expired":   stat_expired,
        "revenue_30d":    revenue_30d,
        "type_choices":   CouponType.choices,
        "status_chips":   status_chips,
        "q": q, "status_filter": status_filter, "channel_filter": channel_filter, "sort": sort,
    })


@admin_required
def coupon_edit(request, pk=None):
    coupon = get_object_or_404(Coupon, pk=pk) if pk else None

    if request.method == "POST":
        from django.utils.dateparse import parse_datetime

        code         = request.POST.get("code", "").strip().upper()
        name         = request.POST.get("name", "").strip()
        description  = request.POST.get("description", "").strip()
        coupon_type  = request.POST.get("coupon_type", CouponType.PERCENT)
        channels_raw = request.POST.getlist("channels")
        channels     = ",".join(channels_raw) if channels_raw else "WEBSITE"

        try:
            discount_value  = float(request.POST.get("discount_value", 0) or 0)
            min_order_value = float(request.POST.get("min_order_value", 0) or 0)
            max_uses        = int(request.POST.get("max_uses", 0) or 0)
            per_user_limit  = int(request.POST.get("per_user_limit", 0) or 0)
        except ValueError:
            discount_value = min_order_value = 0
            max_uses = per_user_limit = 0

        max_disc_raw = request.POST.get("max_discount", "").strip()
        max_discount = float(max_disc_raw) if max_disc_raw else None

        valid_from  = parse_datetime(request.POST.get("valid_from", "").strip()) if request.POST.get("valid_from") else None
        valid_until = parse_datetime(request.POST.get("valid_until", "").strip()) if request.POST.get("valid_until") else None
        is_active   = request.POST.get("is_active") == "1"

        if not code:
            messages.error(request, "Coupon code is required.")
            return redirect(request.path)

        fields = dict(
            code=code, name=name, description=description, coupon_type=coupon_type,
            channels=channels, discount_value=discount_value,
            min_order_value=min_order_value, max_discount=max_discount,
            max_uses=max_uses, per_user_limit=per_user_limit,
            valid_from=valid_from, valid_until=valid_until, is_active=is_active,
        )
        if coupon:
            if Coupon.objects.exclude(pk=coupon.pk).filter(code=code).exists():
                messages.error(request, f'Code "{code}" is already in use.')
                return redirect(request.path)
            for k, v in fields.items():
                setattr(coupon, k, v)
            coupon.save()
            messages.success(request, f'Offer "{code}" updated.')
        else:
            if Coupon.objects.filter(code=code).exists():
                messages.error(request, f'Code "{code}" already exists.')
                return redirect(request.path)
            Coupon.objects.create(**fields)
            messages.success(request, f'Offer "{code}" created.')
        return redirect("/admin-panel/coupons")

    return render(request, "panel/coupon_form.html", {
        "active_path":  "/admin-panel/coupons",
        "coupon":       coupon,
        "type_choices": CouponType.choices,
    })


@admin_required
def coupon_delete(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    code = coupon.code
    coupon.delete()
    messages.success(request, f'Offer "{code}" deleted.')
    return redirect("/admin-panel/coupons")


@admin_required
def coupon_duplicate(request, pk):
    src = get_object_or_404(Coupon, pk=pk)
    new_code = src.code + "_COPY"
    i = 1
    while Coupon.objects.filter(code=new_code).exists():
        new_code = f"{src.code}_COPY{i}"; i += 1
    Coupon.objects.create(
        code=new_code, name=src.name,
        description=src.description, coupon_type=src.coupon_type,
        channels=src.channels, discount_value=src.discount_value,
        min_order_value=src.min_order_value, max_discount=src.max_discount,
        max_uses=src.max_uses, per_user_limit=src.per_user_limit,
        valid_from=src.valid_from, valid_until=src.valid_until,
        is_active=False,
    )
    messages.success(request, f'Offer duplicated as "{new_code}" (inactive by default).')
    return redirect("/admin-panel/coupons")


@admin_required
def coupon_toggle(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.is_active = not coupon.is_active
    coupon.save(update_fields=["is_active"])
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return _JsonResponse({"ok": True, "active": coupon.is_active,
                              "status": coupon.status_tag})
    return redirect("/admin-panel/coupons")


@admin_required
def coupon_detail(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    orders_qs = (Order.objects.filter(coupon=coupon)
                 .select_related("user").order_by("-created_at"))
    order_count  = orders_qs.count()
    total_rev    = orders_qs.aggregate(s=Sum("grand_total"))["s"] or 0
    total_saved  = orders_qs.aggregate(s=Sum("coupon_discount"))["s"] or 0
    conv_rate    = round(coupon.used_count / max(coupon.max_uses, coupon.used_count, 1) * 100) \
                   if coupon.max_uses > 0 else 0

    paginator   = Paginator(orders_qs, 10)
    orders_page = paginator.get_page(request.GET.get("page", 1))

    return render(request, "panel/coupon_detail.html", {
        "active_path":  "/admin-panel/coupons",
        "coupon":       coupon,
        "orders":       orders_page,
        "orders_page":  orders_page,
        "order_count":  order_count,
        "total_rev":    total_rev,
        "total_saved":  total_saved,
        "conv_rate":    conv_rate,
    })


def coupon_validate_api(request):
    """Public AJAX endpoint — validates a coupon code against a cart total."""
    if request.method != "POST":
        return _JsonResponse({"valid": False, "error": "Method not allowed"}, status=405)

    code       = request.POST.get("code", "").strip().upper()
    try:
        cart_total = float(request.POST.get("cart_total", 0) or 0)
    except ValueError:
        cart_total = 0

    try:
        coupon = Coupon.objects.get(code=code)
    except Coupon.DoesNotExist:
        return _JsonResponse({"valid": False, "error": "Invalid coupon code."})

    tag = coupon.status_tag
    if tag == "Inactive":
        return _JsonResponse({"valid": False, "error": "This coupon is inactive."})
    if tag == "Scheduled":
        return _JsonResponse({"valid": False, "error": "This coupon is not active yet."})
    if tag == "Expired":
        return _JsonResponse({"valid": False, "error": "This coupon has expired."})

    if cart_total < coupon.min_order_value:
        return _JsonResponse({
            "valid": False,
            "error": f"Minimum cart value of ₹{coupon.min_order_value:.0f} required."
        })

    discount = coupon.compute_discount(cart_total)
    return _JsonResponse({
        "valid": True,
        "code":       coupon.code,
        "name":       coupon.name or coupon.code,
        "type":       coupon.coupon_type,
        "discount":   discount,
        "free_ship":  coupon.coupon_type == CouponType.FREE_SHIPPING,
        "message":    f'"{coupon.code}" applied! You save ₹{discount:.0f}.',
    })


# ============================================================================ #
#  CMS â€” Content Management System                                              #
# ============================================================================ #

_CMS_PAGES = [
    ("ABOUT_US", "About Us",               "/about-us",         "about"),
    ("TERMS",    "Terms & Conditions",      "/terms",            "terms"),
    ("PRIVACY",  "Privacy Policy",          "/privacy-policy",   "privacy"),
    ("SHIPPING", "Shipping Policy",         "/shipping-policy",  "shipping"),
    ("RETURNS",  "Return & Refund Policy",  "/returns",          "returns"),
    ("CONTACT",  "Contact Us",              "/help-support",     "contact"),
]


@admin_required
def cms_dashboard(request):
    pages = {p.type: p for p in Policy.objects.all()}
    cards = []
    for ptype, label, url, icon in _CMS_PAGES:
        p = pages.get(ptype)
        cards.append({
            "type":       ptype,
            "label":      label,
            "url":        url,
            "icon":       icon,
            "published":  p.is_published if p else False,
            "updated_at": p.updated_at if p else None,
        })
    faq_count      = Faq.objects.count()
    enquiry_unread = Enquiry.objects.filter(is_read=False).count()
    team_count     = TeamMember.objects.count()
    return render(request, "panel/cms_index.html",
                  _ctx("/admin-panel/cms",
                       cards=cards,
                       faq_count=faq_count,
                       enquiry_unread=enquiry_unread,
                       team_count=team_count))


@admin_required
def cms_page_edit(request, page_type):
    page_type = page_type.upper()
    label = dict((k, v) for k, v, *_ in _CMS_PAGES).get(page_type, page_type)

    page, _ = Policy.objects.get_or_create(
        type=page_type,
        defaults={"title": label, "content": ""}
    )

    if request.method == "POST":
        action = request.POST.get("action", "save")
        if page.content:
            CMSRevision.objects.create(
                page_type=page_type,
                title=page.title,
                content=page.content,
                saved_by=request.user,
                note=f"Auto-snapshot before {action}",
            )
        page.title            = request.POST.get("title", page.title).strip()
        page.content          = request.POST.get("content", "")
        page.meta_title       = request.POST.get("meta_title", "").strip()
        page.meta_description = request.POST.get("meta_description", "").strip()
        page.meta_keywords    = request.POST.get("meta_keywords", "").strip()
        page.is_published     = (action == "publish")
        og_file = request.FILES.get("og_image")
        if og_file:
            from core.cloudinary_storage import upload_file as _cl_up
            url, err = _cl_up(og_file.read(), og_file.name, og_file.content_type)
            if url:
                page.og_image = url
            else:
                messages.warning(request, f"OG image upload failed: {err or 'Cloud storage is not configured.'}")
        page.save()
        verb = "Published" if page.is_published else "Saved as draft"
        messages.success(request, f"{verb} â€” {page.title}")
        return redirect(f"/admin-panel/cms/page/{page_type.lower()}")

    revisions = CMSRevision.objects.filter(page_type=page_type)[:10]
    return render(request, "panel/cms_page_edit.html",
                  _ctx("/admin-panel/cms",
                       page=page,
                       page_type=page_type,
                       label=label,
                       revisions=revisions))


@admin_required
def cms_revision_restore(request, pk):
    rev  = get_object_or_404(CMSRevision, pk=pk)
    page, _ = Policy.objects.get_or_create(
        type=rev.page_type, defaults={"title": rev.title, "content": ""}
    )
    CMSRevision.objects.create(
        page_type=rev.page_type, title=page.title, content=page.content,
        saved_by=request.user, note="Auto-snapshot before restore"
    )
    page.title   = rev.title
    page.content = rev.content
    page.save()
    messages.success(request, "Revision restored.")
    return redirect(f"/admin-panel/cms/page/{rev.page_type.lower()}")


# â”€â”€ FAQ Manager â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@admin_required
def cms_faq(request):
    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "add_category":
            name = request.POST.get("name", "").strip()
            if name:
                pos = (FaqCategory.objects.aggregate(m=Max("position"))["m"] or 0) + 1
                FaqCategory.objects.create(name=name, position=pos)
            messages.success(request, "Category added.")
            return redirect("/admin-panel/cms/faq")

        if action == "add_faq":
            cat_id   = request.POST.get("category_id") or None
            question = request.POST.get("question", "").strip()
            answer   = request.POST.get("answer", "").strip()
            if question and answer:
                pos = (Faq.objects.aggregate(m=Max("position"))["m"] or 0) + 1
                Faq.objects.create(
                    category_id=cat_id,
                    question=question,
                    answer=answer,
                    position=pos,
                )
            messages.success(request, "FAQ added.")
            return redirect("/admin-panel/cms/faq")

        if action == "edit_faq":
            faq = get_object_or_404(Faq, pk=request.POST.get("pk"))
            faq.question    = request.POST.get("question", faq.question).strip()
            faq.answer      = request.POST.get("answer", faq.answer).strip()
            faq.category_id = request.POST.get("category_id") or None
            faq.save()
            messages.success(request, "FAQ updated.")
            return redirect("/admin-panel/cms/faq")

        if action == "toggle_faq":
            faq = get_object_or_404(Faq, pk=request.POST.get("pk"))
            from accounts.models import Status as _S
            faq.status = _S.INACTIVE if faq.status == _S.ACTIVE else _S.ACTIVE
            faq.save()
            return redirect("/admin-panel/cms/faq")

        if action == "delete_faq":
            Faq.objects.filter(pk=request.POST.get("pk")).delete()
            repack_positions(Faq)
            messages.success(request, "FAQ deleted.")
            return redirect("/admin-panel/cms/faq")

        if action == "delete_category":
            FaqCategory.objects.filter(pk=request.POST.get("pk")).delete()
            repack_positions(FaqCategory)
            messages.success(request, "Category deleted.")
            return redirect("/admin-panel/cms/faq")

    cats = FaqCategory.objects.filter(is_active=True).prefetch_related("items")
    faqs = Faq.objects.select_related("category").order_by("position")
    return render(request, "panel/cms_faq.html",
                  _ctx("/admin-panel/cms/faq", cats=cats, faqs=faqs))


@admin_required
def cms_faq_reorder(request):
    if request.method == "POST":
        import json as _json
        from django.http import JsonResponse as _JR
        try:
            order = _json.loads(request.body).get("order", [])
            for pos, pk in enumerate(order):
                Faq.objects.filter(pk=pk).update(position=pos)
        except Exception:
            pass
        return _JR({"ok": True})
    return redirect("/admin-panel/cms/faq")


# â”€â”€ About Us â€” rich content + team members â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@admin_required
def cms_about(request):
    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "save_content":
            page, _ = Policy.objects.get_or_create(
                type=PolicyType.ABOUT_US, defaults={"title": "About Us", "content": ""}
            )
            if page.content:
                CMSRevision.objects.create(
                    page_type="ABOUT_US", title=page.title, content=page.content,
                    saved_by=request.user, note="Auto-snapshot"
                )
            page.title            = request.POST.get("title", page.title).strip()
            page.content          = request.POST.get("content", "")
            page.meta_title       = request.POST.get("meta_title", "").strip()
            page.meta_description = request.POST.get("meta_description", "").strip()
            page.is_published     = True
            page.save()
            messages.success(request, "About Us content saved.")
            return redirect("/admin-panel/cms/about")

        if action == "add_member":
            photo_url = ""
            photo = request.FILES.get("photo")
            if photo:
                from core.cloudinary_storage import upload_file as _cl_up
                url, err = _cl_up(photo.read(), photo.name, photo.content_type)
                if url:
                    photo_url = url
                else:
                    messages.warning(request, f"Photo upload failed: {err or 'Cloud storage is not configured.'}")
            pos = (TeamMember.objects.aggregate(m=Max("position"))["m"] or 0) + 1
            TeamMember.objects.create(
                name=request.POST.get("name", "").strip(),
                role=request.POST.get("role", "").strip(),
                bio=request.POST.get("bio", "").strip(),
                photo_url=photo_url,
                position=pos,
            )
            messages.success(request, "Team member added.")
            return redirect("/admin-panel/cms/about")

        if action == "delete_member":
            TeamMember.objects.filter(pk=request.POST.get("pk")).delete()
            repack_positions(TeamMember)
            messages.success(request, "Team member removed.")
            return redirect("/admin-panel/cms/about")

        if action == "toggle_member":
            m = get_object_or_404(TeamMember, pk=request.POST.get("pk"))
            m.is_active = not m.is_active
            m.save(update_fields=["is_active"])
            return redirect("/admin-panel/cms/about")

    page      = Policy.objects.filter(type=PolicyType.ABOUT_US).first()
    team      = TeamMember.objects.all()
    revisions = CMSRevision.objects.filter(page_type="ABOUT_US")[:8]
    return render(request, "panel/cms_about.html",
                  _ctx("/admin-panel/cms/about", page=page, team=team, revisions=revisions))


@admin_required
def cms_team_reorder(request):
    if request.method == "POST":
        import json as _json
        from django.http import JsonResponse as _JR
        try:
            order = _json.loads(request.body).get("order", [])
            for pos, pk in enumerate(order):
                TeamMember.objects.filter(pk=pk).update(position=pos)
        except Exception:
            pass
        return _JR({"ok": True})
    return redirect("/admin-panel/cms/about")


# â”€â”€ Enquiry / Contact inbox â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@admin_required
def cms_enquiries(request):
    status_f   = request.GET.get("status", "")
    qs = Enquiry.objects.select_related("user").prefetch_related("replies")
    if status_f:
        qs = qs.filter(status=status_f)
    paginator  = Paginator(qs, 25)
    page_obj   = paginator.get_page(request.GET.get("page"))
    unread_cnt = Enquiry.objects.filter(is_read=False).count()
    return render(request, "panel/cms_enquiries.html",
                  _ctx("/admin-panel/cms/enquiries",
                       enquiries=page_obj,
                       status_f=status_f,
                       unread_cnt=unread_cnt,
                       EnquiryStatus=EnquiryStatus))


@admin_required
def cms_enquiry_detail(request, pk):
    enq = get_object_or_404(Enquiry, pk=pk)
    if not enq.is_read:
        enq.is_read = True
        enq.save(update_fields=["is_read"])

    if request.method == "POST":
        action = request.POST.get("action", "reply")
        if action == "reply":
            msg = request.POST.get("message", "").strip()
            if msg:
                EnquiryReply.objects.create(enquiry=enq, message=msg, sent_by=request.user)
                enq.status = EnquiryStatus.IN_PROGRESS
                enq.save(update_fields=["status"])
                messages.success(request, "Reply saved.")
        elif action == "close":
            enq.status = EnquiryStatus.CLOSED
            enq.save(update_fields=["status"])
            messages.success(request, "Enquiry closed.")
        elif action == "reopen":
            enq.status = EnquiryStatus.OPEN
            enq.save(update_fields=["status"])
        return redirect(f"/admin-panel/cms/enquiries/{pk}")

    replies = enq.replies.select_related("sent_by").order_by("created_at")
    return render(request, "panel/cms_enquiry_detail.html",
                  _ctx(f"/admin-panel/cms/enquiries",
                       enq=enq,
                       replies=replies,
                       EnquiryStatus=EnquiryStatus))


# ============================================================================ #
#  Testimonials admin                                                           #
# ============================================================================ #

@admin_required
def testimonials_admin(request):
    qs = Testimonial.objects.select_related("user", "order").order_by("-created_at")
    q = request.GET.get("q", "").strip()
    status_f = request.GET.get("status", "")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(comment__icontains=q) | Q(title__icontains=q))
    if status_f:
        qs = qs.filter(approval_status=status_f)
    pending_count = Testimonial.objects.filter(approval_status=ReviewStatus.PENDING).count()
    from django.core.paginator import Paginator
    page = Paginator(qs, 20).get_page(request.GET.get("page", 1))
    return render(request, "panel/testimonials.html",
                  _ctx("/admin-panel/testimonials",
                       testimonials=page, q=q, status_f=status_f,
                       pending_count=pending_count,
                       ReviewStatus=ReviewStatus))


@admin_required
def testimonial_approve(request, pk):
    t = get_object_or_404(Testimonial, pk=pk)
    t.approval_status = ReviewStatus.APPROVED
    t.admin_note = ""
    t.save(update_fields=["approval_status", "admin_note"])
    messages.success(request, f"Testimonial by {t.name} approved.")
    return redirect(request.META.get("HTTP_REFERER", "/admin-panel/testimonials"))


@admin_required
def testimonial_reject(request, pk):
    t = get_object_or_404(Testimonial, pk=pk)
    t.approval_status = ReviewStatus.REJECTED
    t.admin_note = request.POST.get("note", "").strip()
    t.save(update_fields=["approval_status", "admin_note"])
    messages.success(request, f"Testimonial by {t.name} rejected.")
    return redirect(request.META.get("HTTP_REFERER", "/admin-panel/testimonials"))


@admin_required
def testimonial_feature(request, pk):
    t = get_object_or_404(Testimonial, pk=pk)
    t.is_featured = not t.is_featured
    t.save(update_fields=["is_featured"])
    return redirect(request.META.get("HTTP_REFERER", "/admin-panel/testimonials"))


@admin_required
def testimonial_delete_admin(request, pk):
    t = get_object_or_404(Testimonial, pk=pk)
    t.delete()
    messages.success(request, "Testimonial deleted.")
    return redirect("/admin-panel/testimonials")


@admin_required
def order_invoice(request, pk):
    """Standalone print-ready invoice page for an order."""
    import json as _json

    order = get_object_or_404(
        Order.objects.select_related("user", "coupon")
             .prefetch_related("items__product", "items__variant", "items__combo", "refunds"),
        pk=pk,
    )
    cfg = SiteSettings.get()

    try:
        addr = _json.loads(order.shipping_address)
    except Exception:
        addr = {"address": order.shipping_address}

    default_gst = float(cfg.default_gst_rate or 5)
    show_gst    = cfg.show_gst_in_invoice
    prices_incl = cfg.prices_inclusive_of_gst

    # Enrich each item with GST figures
    enriched = []
    total_gst = 0.0
    total_taxable = 0.0
    for it in order.items.all():
        gst_rate = float(it.combo.gst_rate if it.combo else default_gst)
        line = float(it.net_total)
        if prices_incl:
            gst_amt = round(line * gst_rate / (100 + gst_rate), 2)
        else:
            gst_amt = round(line * gst_rate / 100, 2)
        taxable = round(line - gst_amt if prices_incl else line, 2)
        total_gst     += gst_amt
        total_taxable += taxable
        enriched.append({"item": it, "gst_rate": gst_rate, "gst_amt": gst_amt, "taxable": taxable})

    # Split regular vs combo items
    combos_map = {}
    regular_items = []
    for eitem in enriched:
        it = eitem["item"]
        if it.combo_id:
            if it.combo_id not in combos_map:
                combos_map[it.combo_id] = {"combo": it.combo, "items": []}
            combos_map[it.combo_id]["items"].append(eitem)
        else:
            regular_items.append(eitem)

    refunds = list(order.refunds.all())
    total_refunded = sum(r.amount for r in refunds)

    return render(request, "panel/invoice.html", {
        "order":         order,
        "addr":          addr,
        "cfg":           cfg,
        "regular_items": regular_items,
        "combos":        list(combos_map.values()),
        "refunds":       refunds,
        "total_refunded": round(total_refunded, 2),
        "show_gst":       show_gst,
        "prices_incl":    prices_incl,
        "total_gst":      round(total_gst, 2),
        "total_taxable":  round(total_taxable, 2),
        "PaymentMode":    PaymentMode,
        "OrderStatus":    OrderStatus,
        "PaymentStatus":  PaymentStatus,
    })


@admin_required
def admin_change_password(request):
    if request.method == "POST":
        current  = request.POST.get("current_password", "")
        new_pw   = request.POST.get("new_password", "")
        confirm  = request.POST.get("confirm_password", "")

        if not request.user.check_password(current):
            messages.error(request, "Current password is incorrect.")
        elif len(new_pw) < 6:
            messages.error(request, "New password must be at least 6 characters.")
        elif new_pw != confirm:
            messages.error(request, "Passwords do not match.")
        elif new_pw == current:
            messages.error(request, "New password must differ from the current one.")
        else:
            request.user.set_password(new_pw)
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, "Password changed successfully.")
            return redirect("/admin-panel/change-password")

    return render(request, "panel/change_password.html", _ctx("settings"))
